"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:05-12-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import os
from msilib.schema import Font

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.chart import LineChart

class VSCStandardCapacity:

    def __init__(self, excel_file_name, log_directory_path):
        self.excel_file_name = excel_file_name
        self.log_directory_path = log_directory_path
        self.full_excel_file_path = os.path.join(self.log_directory_path, self.excel_file_name + ".xlsx")
        self.log_files = str()
        self.log_list_len = int()
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self.blue_colour = PatternFill(start_color="E7ECFD", end_color="E7ECFD", fill_type="solid")
        self.html_file_path = str()
        self.excel_file_name = "VSC_Standard_capacity"
        self.red_colour = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.red_font = Font(color='00FF0000', bold=True)

    def graph_design(self, sheet):
        chart = LineChart()
        chart.title = "VSC Standard Capacity"
        chart.x_axis.title = 'AU Number'
        chart.y_axis.title = 'Performance [MB/s]'
        chart.legend.position = 'b'

        for col_num in range(1, sheet.max_column + 1):
            data = [cell[0].value for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_num, max_col=col_num)]

            hard_data_numeric = [float(value) for value in data]
            data_ref = Reference(sheet, min_col=col_num, min_row=1, max_col=col_num, max_row=len(hard_data_numeric) + 1)
            chart.add_data(data_ref, titles_from_data=True)

        sheet.add_chart(chart, "G2")
        chart.width = 13
        chart.height = 9

    def graph_update_in_excel_sheet(self):
        try:
            workbook = openpyxl.load_workbook(self.full_excel_file_path)
            sheet = workbook['PW']
            self.graph_design(sheet)
            workbook.save(self.full_excel_file_path)

            workbook = openpyxl.load_workbook(self.full_excel_file_path)
            sheet = workbook['PR']
            self.graph_design(sheet)
            workbook.save(self.full_excel_file_path)
            print("Successfully created the Graphical representation of Performance Values")
        except Exception as err:
            print(f"Error observed: {err}")

    def excel_styling(self):
        """
                Definition to excel styling such as applying thin border, adjust cells width,
                colouring, bold font, merging cells
                """
        def adjust_cell_width(sheet):  # To adjust column width
            max_column_length = 0
            for col in sheet.columns:
                if col[0].column != 1:
                    max_length = max(len(str(cell.value)) for cell in col)
                    max_column_length = max(max_length, max_column_length)
            for col in sheet.columns:
                column = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[column].width = 12 if col[0].column == 1 else max_column_length * 1.0

        def apply_thin_borders(sheet):
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_centre_alignment(sheet):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        try:
            workbook = openpyxl.load_workbook(self.full_excel_file_path)
            sheet_pw = workbook['PW']
            apply_centre_alignment(sheet_pw)
            apply_thin_borders(sheet_pw)
            adjust_cell_width(sheet_pw)
            workbook.save(self.full_excel_file_path)

            workbook = openpyxl.load_workbook(self.full_excel_file_path)
            sheet_pr = workbook['PR']
            apply_centre_alignment(sheet_pr)
            apply_thin_borders(sheet_pr)
            adjust_cell_width(sheet_pr)
            workbook.save(self.full_excel_file_path)
        except Exception as err:
            print("Error observed with excel sytling is {err}")

    def pr_find_occurrences(self,  html_content, vsc_pattern):
        pr_occurrences = []
        pr_start_pattern = f'{vsc_pattern} PR'
        print(f"Start pattern is  {pr_start_pattern}")
        pr_end_pattern = f'{vsc_pattern} Pr - Test Passed!'
        print(f"End pattern is {pr_end_pattern}")

        start_indices = [pos for pos, char in enumerate(html_content) if
                         html_content.startswith(pr_start_pattern, pos)]

        for start_index in start_indices:
            start_pos = start_index + len(pr_start_pattern)
            end_index = html_content.find(pr_end_pattern, start_pos)
            if end_index != -1:
                pr_occurrences.append(html_content[start_pos:end_index].strip())
        return pr_occurrences

    def pr_extract_values(self, html_content_occurance):
        soup = BeautifulSoup(html_content_occurance, 'html.parser')
        pr_performance_value = []
        for label_tag in soup.find_all('span', class_='us', string='VSC Pr [MB/s]'):
            div_tag = label_tag.find_parent('div')
            if div_tag:
                next_div_tag = div_tag.find_next_sibling('div')
                if next_div_tag:
                    value_tag = next_div_tag.find('span', class_='us')
                    if value_tag:
                        int_value = value_tag.text.strip()
                        if int_value is not None and int_value != "":
                            try:
                                int_value = round(float(int_value), 8)
                                pr_performance_value.append(int_value)
                            except ValueError:
                                print(f"Error: Unable to convert '{int_value}' to float.")
                                pr_performance_value.append(float(0))  # Appending 0 to empty values
                        else:
                            pr_performance_value.append(float(0))
        return pr_performance_value

    def pw_find_occurrences(self,  html_content, vsc_pattern):
        pw_occurrences = []
        pw_start_pattern = f'{vsc_pattern} Pw'
        print(f"Start pattern is  {pw_start_pattern}")
        pw_end_pattern = f'{vsc_pattern} Pw - Test Passed!'
        print(f"End pattern is {pw_end_pattern}")

        start_indices = [pos for pos, char in enumerate(html_content) if
                         html_content.startswith(pw_start_pattern, pos)]

        for start_index in start_indices:
            start_pos = start_index + len(pw_start_pattern)
            end_index = html_content.find(pw_end_pattern, start_pos)
            if end_index != -1:
                pw_occurrences.append(html_content[start_pos:end_index].strip())
        return pw_occurrences

    def pw_extract_values(self, html_content_occurance):
        soup = BeautifulSoup(html_content_occurance, 'html.parser')
        pw_performance_value = []
        for label_tag in soup.find_all('span', class_='us', string='VSC Pw [MB/s]'):
            div_tag = label_tag.find_parent('div')
            if div_tag:
                next_div_tag = div_tag.find_next_sibling('div')
                if next_div_tag:
                    value_tag = next_div_tag.find('span', class_='us')
                    if value_tag:
                        int_value = value_tag.text.strip()
                        if int_value is not None and int_value != "":  # Check if int_value is not None or an empty string
                            try:
                                int_value = round(float(int_value), 8)
                                pw_performance_value.append(int_value)
                            except ValueError:
                                print(f"Error: Unable to convert '{int_value}' to float.")
                                pw_performance_value.append(float(0))
                        else:
                            pw_performance_value.append(float(0))
        return pw_performance_value

    def extracting_values_from_log_files(self):
        self.html_file_path = os.path.join(self.log_directory_path, self.log_files)
        print(self.html_file_path)

        def process_sheet(sheet_name, find_occurrences_func, extract_values_func):
            result_dict = {}
            for vsc_pattern in vsc_patterns:
                occurrences = find_occurrences_func(html_content, vsc_pattern)
                result_dict[f'{vsc_pattern}_values'] = []
                for occurrence in occurrences:
                    performance_value = extract_values_func(occurrence)
                    result_dict[f'{vsc_pattern}_values'].extend(performance_value)
            for string, value in result_dict.items():
                print(f"Extracted {sheet_name} values for {string}:", value)
            return result_dict

        with open(self.html_file_path, 'r') as file:
            html_content = file.read()
        vsc_patterns = ['VSC6', 'VSC10', 'VSC30', 'VSC60', 'VSC90']
        pw_result_dict = process_sheet('PW', self.pw_find_occurrences, self.pw_extract_values)
        pr_result_dict = process_sheet('PR', self.pr_find_occurrences, self.pr_extract_values)

        try:
            workbook = openpyxl.load_workbook(self.full_excel_file_path)

            def update_sheet(sheet, column_names, result_dict):
                for string, indices in result_dict.items():
                    st_value = string.split("_")[0]
                    for col_index, col_name in enumerate(column_names, start=1):
                        if st_value == col_name:
                            specified_value = col_name[3:]
                            for i, value in enumerate(indices, start=2):
                                if value == float(0.0):
                                    sheet.cell(row=i, column=col_index).font = self.red_font
                                elif value < float(specified_value):
                                    sheet.cell(row=i, column=col_index, value=value)
                                    sheet.cell(row=i, column=col_index).font = self.red_font
                                sheet.cell(row=i, column=col_index, value=value)

            update_sheet(workbook['PW'], [cell.value for cell in workbook['PW'][1]], pw_result_dict)
            update_sheet(workbook['PR'], [cell.value for cell in workbook['PR'][1]], pr_result_dict)

            workbook.save(self.full_excel_file_path)
            print("Successfully updated and saved the Excel file with 'PW' and 'PR' sheets.")
        except Exception as err:
            print(f"Error occurred while updating and saving the Excel file: {err}")

    def creating_excel_file_with_header(self):
        try:
            self.log_files = os.listdir(self.log_directory_path)
            for file in self.log_files:
                if file.endswith('.htm'):
                    self.log_files = file
            self.html_file_path = os.path.join(self.log_directory_path, self.log_files)
            print(self.html_file_path)

            with open(self.html_file_path, 'r') as file:
                html_content = file.read()
                if html_content.__contains__("VSC90"):
                    headers = ["VSC6", "VSC10", "VSC30", "VSC60", "VSC90"]
                    print(f"Currently Parsing V90 Log file {headers}")
                else:
                    headers = ["VSC6", "VSC10", "VSC30", "VSC60"]
                    print(f"Currently Parsing V60 Log file {headers}")
            self.log_list_len = len(self.log_files)
            print(f"List of Log files are {self.log_files}")
            print(f"Total no. of log files are  {self.log_list_len}")

        except FileNotFoundError:
            print("File are not found or provided path is incorrect")
        except Exception as err:
            print(f"Error occurred while iteration log files is {err}")

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "PW"
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = self.blue_colour

            sheet = workbook.create_sheet(title="PR")
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = self.blue_colour

            workbook.save(self.full_excel_file_path)
            print("Successfully created the VSC_Standard_capacity Excel Template")
        except Exception as err:
            print(f"Error observed while creating the {self.excel_file_name} file")

def input_from_cmd_line():
    print("=" * 55 + "VSC Standard Capacity Parsing Started" + "=" * 55)
    print("Enter the file name to get VSC Standard capacity as: VSC_UHS_II")
    file_name = input("Enter the VSC Standard capacity Parsing file name: ")
    log_path = input("Enter the path of log files: ")
    return file_name, log_path


if __name__ == "__main__":
    excel_file_name, log_directory_path = input_from_cmd_line()
    print(f"Entered VSC Standard capacity file name is {excel_file_name}")
    print(f"Entered log files directory is {log_directory_path}")
    VSC_SC = VSCStandardCapacity(excel_file_name, log_directory_path)
    VSC_SC.creating_excel_file_with_header()
    VSC_SC.extracting_values_from_log_files()
    VSC_SC.excel_styling()
    VSC_SC.graph_update_in_excel_sheet()
