"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:04-12-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class CDMSequentialParser:

    def __init__(self, excel_file_name, log_directory_path):
        self.excel_file_name = excel_file_name
        self.log_directory_path = log_directory_path
        self.full_excel_file_path = os.path.join(self.log_directory_path, self.excel_file_name + ".xlsx")
        self.cdm_logs_list = []
        self.log_list = []
        self.log_list_length = int()
        self.row_name_list = []
        self.start_row_index = 5
        self.search_string = ["Sequential Read (Q= 32,T= 1)", "Sequential Write (Q= 32,T= 1)"]
        self.col_num = [2, 3]
        self.red_colour = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self.red_font = Font(color='00FF0000', bold=True)
        self.cdm_version = "CrystalDiskMark"
        self.test_data = "Test"
        self.grey_colour = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
        self.blue_colour = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        self.green_colour = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    def excel_styling(self):
        def adjust_cell_width(sheet):
            for col in sheet.columns:
                max_length = 0
                column = openpyxl.utils.get_column_letter(col[0].column)
                for cells in col:
                    try:
                        if len(str(cells.value)) > max_length:
                            max_length = len(str(cells.value))
                    except Exception as e:
                        print(f"{e}")
                        pass
                adjusted_width = (max_length + 2) * 0.9
                sheet.column_dimensions[column].width = adjusted_width

        def apply_thin_borders(sheet):
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_centre_alignment(sheet):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        try:
            wb = openpyxl.load_workbook(self.full_excel_file_path)
            sheet = wb.active
            adjust_cell_width(sheet)
            apply_thin_borders(sheet)
            apply_centre_alignment(sheet)

            sheet.merge_cells('A1:C1')
            sheet.merge_cells('B2:C2')
            sheet.merge_cells('B3:C3')

            wb.save(self.full_excel_file_path)
            print(f"Excel styling are updated in {self.full_excel_file_path} in the Excel file.")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred during excel styling: {e}")

    def updating_final_values(self):
        try:
            wb = openpyxl.load_workbook(self.full_excel_file_path)
            sheet = wb.active
            cell_indices = [4 + self.log_list_length + i for i in range(1, 4)]

            for i, label in enumerate(["Average", "Max Value", "Min Value"]):
                cell_index = cell_indices[i]
                sheet[f"A{cell_index}"] = label
                sheet[f"A{cell_index}"].font = Font(bold=True)
                sheet[f"A{cell_index}"].fill = self.grey_colour

                if i == 0:
                    read_formula = f"=AVERAGE(B5:B{cell_index - 1})"
                    write_formula = f"=AVERAGE(C5:C{cell_index - 1})"
                elif i == 1:
                    read_formula = f"=MAX(B5:B{cell_index - 1})"
                    write_formula = f"=MAX(C5:C{cell_index - 1})"
                else:
                    read_formula = f"=MIN(B5:B{cell_index - 1})"
                    write_formula = f"=MIN(C5:C{cell_index - 1})"

                sheet[f"B{cell_index}"] = read_formula
                sheet[f"B{cell_index}"].fill = self.grey_colour

                sheet[f"C{cell_index}"] = write_formula
                sheet[f"C{cell_index}"].fill = self.grey_colour

            wb.save(self.full_excel_file_path)
            print(f"Final values are updated in {self.full_excel_file_path} in the Excel file.")

        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred during updating final values: {e}")

    def extracting_wr_values_from_log_file(self):
        row = 5
        for txt_file, column_name in zip(self.log_list, self.row_name_list):
            txt_file_full_path = os.path.join(self.log_directory_path,txt_file)
            print(f"Reading txt format log file is {txt_file_full_path}")
            print(f"Updating current card values in {column_name}")
            extracted_values_list = []
            with open(txt_file_full_path, 'r', encoding='utf-8', errors='ignore') as file:
                txt_content = file.read()
                if self.cdm_version in txt_content:
                    lines = txt_content.split('\n')
                    for line in lines:
                        if self.cdm_version in line:
                            cdm_vr = line.split(" ")[1]
                            print(f"CDM Version is {cdm_vr}")
                if self.test_data in txt_content:
                    lines = txt_content.split('\n')
                    for line in lines:
                        if self.test_data in line:
                            test_data = str(line.split(":")[1]).strip()
                            test_data_bytes = test_data.split(" ")[0]
                            test_data_unit = test_data.split(" ")[1]
                            td = test_data_bytes+" "+test_data_unit
                            td = str(td).strip()
                            if td == r"1024 MiB":
                                test_range = "1GB"
                            print(f"Test Data is {test_range}")
                for string in self.search_string:
                    if string in txt_content:
                        lines = txt_content.split('\n')
                        for line in lines:
                            if string in line:
                                print(line)
                                value = (str(line.split(":")[1]).strip())
                                value = value.split(" ")[0]
                                extracted_values_list.append(value)
                                break
                try:
                    wb = openpyxl.load_workbook(self.full_excel_file_path)
                    sheet = wb.active

                    sheet['B2'] = cdm_vr
                    sheet['B3'] = test_range

                    for i, val in zip(self.col_num, extracted_values_list):
                        current_col = openpyxl.utils.get_column_letter(i)
                        if val == '' or val == "nan" or val == " " or val is None:
                            print(f"No values encountered.Please check Log {column_name} files")
                            sheet[f"{current_col}{row}"].fill = self.red_colour
                        else:
                            if val == "MB/s":
                                sheet[f"{current_col}{row}"].value = ' '
                                print(f"No values encountered.Please check Log {column_name} files")
                                sheet[f"{current_col}{row}"].fill = self.red_colour
                            else:
                                sheet[f"{current_col}{row}"] = float(val)
                    row += 1
                    wb.save(self.full_excel_file_path)
                    print(f"Extracted values are updated in {self.full_excel_file_path} in the Excel file.")
                except FileNotFoundError:
                    print("File not found or paths are incorrect")
                except Exception as e:
                    print(f"An error occurred during updating sample number column: {e}")

    def update_sample_num_column(self):
        try:
            wb = openpyxl.load_workbook(self.full_excel_file_path)
            sheet = wb.active
            for log_file_name in self.log_list:
                row_name = int(log_file_name.split(".")[0])
                row_name = f"Card-{row_name}"
                self.row_name_list.append(row_name)
                sheet["A" + f"{self.start_row_index}"].value = row_name
                self.start_row_index += 1
            wb.save(self.full_excel_file_path)
            print(f"{self.log_list_length} samples created from row '{self.start_row_index}' in the Excel file.")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred during updating sample number column: {e}")

    def iterating_over_log_directory(self):
        """
        Defintion to get the list of log_files_names and total no. of log file
        """
        try:
            self.cdm_logs_list = os.listdir(self.log_directory_path )
            self.log_list = [file for file in self.cdm_logs_list if file.endswith('.txt')]
            self.log_list = sorted(self.log_list, key=lambda x: int(x.split('.')[0]))
            self.log_list_length = len(self.log_list)
            print(f"Total number of log files in a directory: {self.log_list_length}")
            print(f"List of log files are {self.log_list}")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as err:
            print(f"Occurred exception while iterating log directory is {err}")

    def creating_CDM_sequential_excel_template(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            sheet['A1'] = "CDM Performance"
            sheet['A1'].fill = self.green_colour
            sheet['A2'] = "CDM Version"
            sheet['A2'].fill = self.blue_colour
            sheet['A3'] = "Data Range"
            sheet['A3'].fill = self.blue_colour
            sheet['A4'] = "Sample No"
            sheet['A4'].fill = self.blue_colour
            sheet['B4'] = "Read (MB/s)"
            sheet['B4'].fill = self.blue_colour
            sheet['C4'] = "Write (MB/s)"
            sheet['C4'].fill = self.blue_colour

            def bold_font_cells(sheet, bold_cells):
                for cells in bold_cells:
                    sheet[cells].font = openpyxl.styles.Font(bold=True)
            bold_cells = ['A1','A2','A3','A4','B4','C4']
            bold_font_cells(sheet, bold_cells)


            workbook.save(self.full_excel_file_path)
            print("Successfully created the CDM Performance Excel Template")
        except Exception as err:
            print(f"Error observed while creating the {self.full_excel_file_path} file is {err}")

def input_from_cmd_line():
    print("=" * 55 + "CDM Sequential Parsing Started" + "=" * 55)
    print("Enter the CDM Sequential Parsing file name as : CDM_Performance")
    file_name = input("Enter the CDM Sequential Parsing file name: ")
    log_path = input("Enter the path of log files: ")
    return file_name, log_path

if __name__ == "__main__":
    excel_file_name, log_directory_path = input_from_cmd_line()
    print(f"Entered A1 CDM file name is {excel_file_name}")
    print(f"Entered log files directory is {log_directory_path}")
    CDMseq = CDMSequentialParser(excel_file_name, log_directory_path)
    CDMseq.creating_CDM_sequential_excel_template()
    CDMseq.iterating_over_log_directory()
    CDMseq.update_sample_num_column()
    CDMseq.extracting_wr_values_from_log_file()
    CDMseq.updating_final_values()
    CDMseq.excel_styling()