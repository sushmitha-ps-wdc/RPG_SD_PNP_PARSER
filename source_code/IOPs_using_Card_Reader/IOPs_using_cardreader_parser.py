"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:29-11-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles import Border, Side
import csv


class IOPSUsingCardReaderParser:

    def __init__(self, file_name, log_directory_path):
        self.file_name = file_name
        self.log_directory_path = log_directory_path
        self.excel_file_name = os.path.join(self.log_directory_path, self.file_name + ".xlsx")
        self.log_files_list = []
        self.log_list_len = int()
        self.column_name_list = []
        self.strings_to_find = ["RRead", "SRead", "SWrite", "RWrite"]
        self.search_strings = ["RRead 4K", "SRead 64K", "SWrite 64K", "RWrite 4K"]
        self.yellow_colour = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.light_orange_colour = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        self.red_colour = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.peach_color = PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid")
        self.grey_color = PatternFill(start_color="DDD9C4", end_color="DDD9C4", fill_type="solid")

        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))

    def styling_excel(self):
        def adjust_cell_width(sheet):
            for col in sheet.columns:
                max_length = 0
                column = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                        print(f"Error: {e}")
                        pass
                adjusted_width = (max_length + 1) * 0.8
                sheet.column_dimensions[column].width = adjusted_width

        def apply_thin_borders(sheet):
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_centre_alignment(sheet):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        try:
            workbook = openpyxl.load_workbook(self.excel_file_name)
            sheet = workbook.active
            adjust_cell_width(sheet)
            apply_thin_borders(sheet)
            apply_centre_alignment(sheet)

            def bold_font_cells(sheet, bold_cells):
                for cells in bold_cells:
                    sheet[cells].font = openpyxl.styles.Font(bold=True, color="333399")  # Blue font
                for cells in bold_cells:
                    sheet[cells].fill = self.grey_color
            bold_font_cell_list = ['B1','A2','A3','A4','A5','A6','A10','A11','A12','A13','A14','B7','B9']
            bold_font_cells(sheet, bold_font_cell_list)

            last_column_char_to_merge = chr(65 + self.log_list_len % 26)
            # To get Last column index for merging cells
            try:
                merge_ranges = [
                    f"B1:{last_column_char_to_merge}1",
                    f"B7:{last_column_char_to_merge}7",
                    f"B9:{last_column_char_to_merge}9"
                ]
                for merge_range in merge_ranges:
                    sheet.merge_cells(merge_range)

                def fill_color_in_range(sheet, cell_range, color):
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    for row in sheet[cell_range]:
                        for cell in row:
                            cell.fill = fill

                for cell_range in merge_ranges:
                    fill_color_in_range(sheet, cell_range, "FABF8F") # peach colour

            except ValueError as ve:
                print(f"Error while merging cells: {ve}")
            finally:
                workbook.save(self.excel_file_name)
        except Exception as e:
            print(f"Exception occurred while updating Excel styling: {e}")

    def extracting_values_from_csv_log_files(self):
        column_num = 2
        for csv_file, column_name in zip(self.log_files_list, self.column_name_list):
            csv_file_full_path = os.path.join(self.log_directory_path, csv_file)
            print("\n\r")
            print(f"Reading csv log file is {csv_file_full_path}")
            print(f"Updating current card values in {column_name}")

            extracted_values = {string: None for string in self.search_strings}
            found_indices = {string: False for string in self.search_strings}

            with open(csv_file_full_path, 'r') as file:
                csv_reader = csv.reader(file)
                next(csv_reader)
                for row in csv_reader:
                    if len(row) >= 3:
                        for search_string in self.search_strings:
                            if search_string == "RRead 4K" or search_string == "RWrite 4K":
                                if not found_indices[search_string] and search_string in row[2] and row[6].strip() != '':
                                    found_indices[search_string] = True
                                    extracted_values[search_string] = float(row[6]) if row[6].strip() != '' else 0
                                    break
                            elif search_string == "SRead 64K" or search_string == "SWrite 64K":
                                if not found_indices[search_string] and search_string in row[2] and row[9].strip() != '':
                                    found_indices[search_string] = True
                                    extracted_values[search_string] = float(row[9]) if row[9].strip() != '' else 0
                                    break

            print(extracted_values)
            row_indices = {"RRead 4K": [3],
                           "SRead 64K": [4],
                           "SWrite 64K": [5],
                           "RWrite 4K": [6]
                           }
            print(f"row_indices{row_indices}")

            try:
                wb = openpyxl.load_workbook(self.excel_file_name)
                sheet = wb.active

                for string, indices in row_indices.items():
                    values = extracted_values[string]
                    if isinstance(values, list):
                        for i, val in zip(indices, values):
                            if val is None:
                                sheet.cell(row=i, column=column_num, value=float(0))
                            else:
                                sheet.cell(row=i, column=column_num, value=float(val))
                    else:
                        if values is None or values == 0:
                            sheet.cell(row=indices[0], column=column_num, value=float(0))
                            cell = sheet.cell(row=indices[0], column=column_num)
                            cell.font = Font(bold=True, color="FF0000")
                        else:
                            sheet.cell(row=indices[0], column=column_num, value=float(values))

                        if string == "RRead 4K" and values is not None and float(values) <= 1500:
                            cell = sheet.cell(row=indices[0], column=column_num)
                            cell.font = Font(bold=True, color="FF0000")
                        elif string == "RWrite 4K" and values is not None and float(values) <= 500:
                            cell = sheet.cell(row=indices[0], column=column_num)
                            cell.font = Font(bold=True, color="FF0000")

                wb.save(self.excel_file_name)
            except Exception as err:
                print(f"Exception occurred while updating values to excel: {err}")
            column_num += 1

    def creating_sample_column(self):
        try:
            column_num = 2
            wb = openpyxl.load_workbook(self.excel_file_name)
            sheet = wb.active
            for col, log_name in zip((range(column_num, column_num + self.log_list_len)), self.log_files_list):
                current_col = openpyxl.utils.get_column_letter(col)
                log_name = int(log_name.split(".")[0])
                column_name = f"Card-{log_name}"
                self.column_name_list.append(column_name)
                sheet[f"{current_col}2"].value = column_name
                sheet[f"{current_col}2"].font = Font(bold=True)
                sheet[f"{current_col}2"].fill = self.yellow_colour

                formula_cell1 = f"{current_col}{8}"
                data_cell1 = f"{current_col}3"
                sheet[formula_cell1].value = f"=(1000/{data_cell1})-0.17"

                sheet[f"{current_col}10"].value = sheet[f"{current_col}2"].value  # card name
                sheet[f"{current_col}10"].font = Font(bold=True)
                sheet[f"{current_col}10"].fill = self.yellow_colour

                formula_cell2 = f"{current_col}{11}"
                data_cell2 = f"{current_col}8"
                sheet[formula_cell2].value = f"=(1/{data_cell2})* 1000"

                sheet[f"{current_col}12"].value = f"=({current_col}4)"
                sheet[f"{current_col}13"].value = f"=({current_col}5)"
                sheet[f"{current_col}14"].value = f"=({current_col}6)"

            col_index_required = column_num + self.log_list_len
            col_index_required = openpyxl.utils.get_column_letter(col_index_required)

            sheet[f"{col_index_required}10"].value = "Required"
            sheet[f"{col_index_required}10"].font = Font(bold=True,color="008000")
            sheet[f"{col_index_required}10"].fill = self.yellow_colour

            required_row_index_list = [11,12,13,14]
            required_row_index_value = [1500,10,10,500]
            for row_index, row_value in zip(required_row_index_list, required_row_index_value):
                sheet[f"{col_index_required}{row_index}"].value = row_value
                sheet[f"{col_index_required}{row_index}"].font = Font(bold=True, color="008000") #green font

            col_index_ratio = column_num + self.log_list_len + 1
            col_index_ratio = openpyxl.utils.get_column_letter(col_index_ratio)

            sheet[f"{col_index_ratio}10"].value = "Ratio"
            sheet[f"{col_index_ratio}10"].font = Font(bold=True)
            sheet[f"{col_index_ratio}10"].fill = self.light_orange_colour

            ratio_row_index_list = [11, 12, 13, 14]
            for ri in ratio_row_index_list:
                ratio_formula = f"=1-({col_index_required}{ri}/MIN(B{ri}:D{ri}))"
                sheet[f"{col_index_ratio}{ri}"].value = ratio_formula

            for per in ratio_row_index_list:
                sheet[f"{col_index_ratio}{per}"].number_format = '0.00%'
                sheet[f"{col_index_ratio}{per}"].font = Font(bold=True, color="333399")
                sheet[f"{col_index_ratio}{per}"].fill = self.peach_color

            wb.save(self.excel_file_name)
            print(f"Successfully update the column name and formulas to file {self.excel_file_name}")
        except Exception as err:
            print(f"Error occurred while creating sample column and updating formulas is {err}")

    def iterating_over_log_directory(self):
        try:
            self.log_files_list = os.listdir(self.log_directory_path)
            self.log_files_list = [file for file in self.log_files_list if file.endswith('.csv')]
            self.log_files_list = sorted(self.log_files_list, key=lambda x: int(x.split(".")[0]))
            self.log_list_len = len(self.log_files_list)
            print(f"List of Log files are {self.log_files_list}")
            print(f"Total no. of log files are  {self.log_list_len}")
        except FileNotFoundError:
            print("File are not found or provided path is incorrect")
        except Exception as err:
            print(f"Error occurred while iteration log files is {err}")

    def creating_iops_file(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet['B1'] = "IOMeter"
            sheet['A2'] = sheet['A10'] = "Operations"
            sheet['A3'] = sheet['A11'] = "RRead 4K [IOPs]"
            sheet['A4'] = sheet['A12'] = "SRead 64K [MBps]"
            sheet['A5'] = sheet['A13'] = "SWrite 64K [MBps]"
            sheet['A6'] = sheet['A14'] = "RWrite 4K [IOPs]"
            sheet['B7'] = "HTAT"
            sheet['B9'] = "Final Results without HTAT"
            workbook.save(self.excel_file_name)
            print(f"Successfully created the IOPs Excel Template {self.excel_file_name}")
        except Exception as err:
            print(f"Exception occurred while creating the IOPS file: {err}")

def input_from_cmd_line():
    print("=" * 55 + "IOPS USING CARD READER Parsing Started" + "=" * 55)
    print("Enter following file name for IOPS as: 'IOPs'")
    file_name = input("Enter the file name: ")
    log_directory_path = input("Enter the path of log files: ")
    return file_name, log_directory_path


if __name__ == "__main__":
    file_name, log_directory_path = input_from_cmd_line()
    print(f"Entered file name is {file_name}")
    print(f"Entered log files directory is {log_directory_path}")
    iops = IOPSUsingCardReaderParser(file_name, log_directory_path)
    iops.creating_iops_file()
    iops.iterating_over_log_directory()
    iops.creating_sample_column()
    iops.extracting_values_from_csv_log_files()
    iops.styling_excel()
