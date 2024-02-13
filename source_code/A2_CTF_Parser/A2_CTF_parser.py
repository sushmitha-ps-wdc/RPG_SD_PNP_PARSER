"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:28-11-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment, Border, Side


class A2CTFParser:

    def __init__(self, excel_file_name, log_directory_path):
        self.excel_file_name = excel_file_name
        self.log_directory_path = log_directory_path
        self.full_excel_file_path = os.path.join(self.log_directory_path,self.excel_file_name + ".xlsx")
        self.log_files_list = []
        self.log_list_len = int()
        self.column_name_list = []
        self.extracted_values = []
        self.search_string = ["Random Read IOPS for IO size 4K", "Random Write IOPS for IO size 4K"]
        self.extracted_values_list = []
        self.rows_num = [3, 4]
        self.red_font = Font(color='00FF0000', bold=True)
        self.red_colour = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self.dark_yellow_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.grey_color = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self.light_green_color = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


    def excel_styling(self):
        def fill_first_row_with_color():
            for cell in sheet[1]:
                if cell.value is not None:
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type='solid')

        def adjust_cell_width():
            for col in sheet.columns:
                max_length = 0
                column = openpyxl.utils.get_column_letter(col[0].column)
                for cells in col:
                    try:
                        if len(str(cells.value)) > max_length:
                            max_length = len(str(cells.value))
                    except Exception as e:
                        print(f"{e}")
                        pass
                adjusted_width = (max_length + 2) * 0.9
                sheet.column_dimensions[column].width = adjusted_width

        def apply_thin_borders():
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_centre_alignment():
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        try:
            workbook = openpyxl.load_workbook(self.full_excel_file_path)
            sheet = workbook.active
            fill_first_row_with_color()
            adjust_cell_width()
            apply_thin_borders()
            apply_centre_alignment()
            max_column = sheet.max_column
            current_col = openpyxl.utils.get_column_letter(max_column)
            sheet.merge_cells(f"A1:{current_col}1")

            workbook.save(self.full_excel_file_path)
        except Exception as err:
            print(f"Error occurred while updating excel sheet format is {err}")

    def extracting_values_from_log_files(self):
        column_num  = 3
        for txt_file, column_name in zip(self.log_files_list, self.column_name_list):
            txt_file_full_path = os.path.join(self.log_directory_path,txt_file)
            print(f"Reading txt format log file is {txt_file_full_path}")
            print(f"Updating current card values in {column_name}")
            with open(txt_file_full_path, 'r', encoding='utf-8', errors='ignore') as file:
                self.extracted_values_list = []
                txt_content = file.read()
                for string in self.search_string:
                    if string in txt_content:
                        lines = txt_content.split('\n')
                        for line in lines:
                            if string in line:
                                value = (str(line.split("=")[1]).strip())
                                self.extracted_values_list.append(value)
                                break
            print(self.extracted_values_list)
            try:
                wb = openpyxl.load_workbook(self.full_excel_file_path)
                sheet = wb.active
                for i, val in zip(self.rows_num,self.extracted_values_list):
                    current_col = openpyxl.utils.get_column_letter(column_num)
                    if val == '' or val == "nan" or val == " " or val is None:
                        print(f"No values encountered.Please check Log {column_name} files")
                        sheet[f"{current_col}{i}"].fill = self.red_colour
                    else:
                        if val == "IOPS]":
                            sheet[f"{current_col}{i}"].value = ' '
                            print(f"No values encountered.Please check Log {column_name} files")
                            sheet[f"{current_col}{i}"].fill = self.red_colour
                        else:
                            sheet[f"{current_col}{i}"] = float(val)
                            thresholds = [4000, 2000]

                            for threshold, target_column in zip(thresholds, range(3, 5)):
                                if i == target_column:
                                    if float(val) >= threshold:
                                        pass
                                    else:
                                        print(
                                            f"Actual value is not expected to Specified Value in log file {column_name}")
                                        sheet[f"{current_col}{i}"].font = self.red_font
                print("Successfully updated the extracted values to A2 CTF File")
            except Exception as err:
                print(f"Error occurred while updating values is {err}")
            column_num += 1
            wb.save(self.full_excel_file_path)

    def creating_sample_column(self):
        try:
            column_num = 3
            wb = openpyxl.load_workbook(self.full_excel_file_path)
            sheet = wb.active
            for col,log_name in zip((range(column_num, column_num + self.log_list_len)), self.log_files_list):
                current_col = openpyxl.utils.get_column_letter(col)
                log_name = int(log_name.split(".")[0])
                column_name = f"Card-{log_name}"
                self.column_name_list.append(column_name)
                sheet[f"{current_col}2"].value = column_name
                sheet[f"{current_col}2"].font = Font(bold =True)
                sheet[f"{current_col}2"].fill = self.dark_yellow_color

            wb.save(self.full_excel_file_path)
            print("Successfully update the column name as log file names")
        except Exception as err:
            print(f"Error occurred while creating sample column is {err}")

    def iterating_over_log_directory(self):
        try:
            self.log_files_list = os.listdir(self.log_directory_path)
            self.log_files_list = [file for file in self.log_files_list if file.endswith('.log')]
            self.log_files_list = sorted(self.log_files_list, key = lambda x: int(x.split(".")[0]))
            self.log_list_len = len(self.log_files_list)
            print(f"List of Log files are {self.log_files_list}")
            print(f"Total no. of log files are  {self.log_list_len}")
        except FileNotFoundError:
            print("File are not found or provided path is incorrect")
        except Exception as err:
            print(f"Error occured while iteration log files is {err}")

    def creating_A2_CTF_excel_template(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet['A1'] = "A2 Performance using CTF "
            sheet['A1'].font = Font(bold=True)
            sheet['A2'] = "Operation"
            sheet['A2'].font = Font(bold=True, color="0000FF")
            sheet['A2'].fill = self.grey_color

            sheet['B2'] = "Specified"
            sheet['B2'].font = Font(bold=True)
            sheet['B2'].fill = self.light_green_color

            sheet['A3'] = "IOPS_RND_R_4KB_1GB[IOPs]"
            sheet['A3'].fill = self.grey_color
            sheet['A3'].font = Font(bold=True, color="0000FF")

            sheet['A4'] = "IOPS_RND_W_4KB_1GB[IOPs]"
            sheet['A4'].fill = self.grey_color
            sheet['A4'].font = Font(bold=True, color="0000FF")

            sheet['B3'] = float("4000")
            sheet['B4'] = float("2000")
            workbook.save(self.full_excel_file_path)
            print("Successfully created the A2 CTF Excel Template")
        except Exception as err:
            print(f"Error observed while creating the {self.excel_file_name} file")

def input_from_cmd_line():
    print("=" * 55 + "A2 CTF Parsing Started" + "=" * 55)
    file_name = input("Enter the A2 CTF Parsing file name: ")
    log_path = input("Enter the path of log files: ")
    return file_name, log_path

if __name__ == "__main__":
    excel_file_name, log_directory_path = input_from_cmd_line()
    print(f"Entered A2 CTF file name is {excel_file_name}")
    print(f"Entered log files directory is {log_directory_path}")
    A2_CTF = A2CTFParser(excel_file_name, log_directory_path)
    A2_CTF.creating_A2_CTF_excel_template()
    A2_CTF.iterating_over_log_directory()
    A2_CTF.creating_sample_column()
    A2_CTF.extracting_values_from_log_files()
    A2_CTF.excel_styling()