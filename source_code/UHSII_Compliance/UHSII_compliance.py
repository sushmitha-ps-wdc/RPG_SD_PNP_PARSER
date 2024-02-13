"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:8-12-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import os
import re

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
import pandas as pd


class UHSIICompliance:

    def __init__(self,excel_file_name, log_directory_path):
        self.excel_file_name = excel_file_name
        self.log_directory_path = log_directory_path
        self.excel_file_name = os.path.join(self.log_directory_path, self.excel_file_name + ".xlsx")
        self.log_files_list = []
        self.log_list_len = int()
        self.column_name_list = []
        self.blue_colour = PatternFill(start_color="E7ECFD", end_color="E7ECFD", fill_type="solid")
        self.orange_colour = PatternFill(start_color="FFD27F", end_color="FFD27F", fill_type="solid")
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))


    def excel_styling(self):
        """
        Definition to excel styling such as applying thin border, adjust cells width,
        colouring, bold font, merging cells
        """

        def adjust_cell_width(sheet):
            max_column_length = {}

            for col in sheet.columns:
                column_index = col[0].column
                max_length = max(len(str(cell.value)) for cell in col)
                max_column_length[column_index] = max_length

            for col in sheet.columns:
                column_index = col[0].column
                if column_index == 2:
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = max_column_length[
                                                                                                        column_index] * 0.9
                else:
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = max_column_length[column_index] * 1.0

        def apply_thin_borders(sheet):
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_alignment(sheet):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for col_num, cell in enumerate(row, start=1):
                    if col_num == 2:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

        def fill_first_column(sheet):
            start_row = 2

            for row_num, row in enumerate(
                    sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=1),
                    start=start_row):
                for cell in row:
                    cell.value = row_num - 1

        try:
            workbook = openpyxl.load_workbook(self.excel_file_name)
            sheet = workbook.active
            fill_first_column(sheet)
            apply_alignment(sheet)
            apply_thin_borders(sheet)
            adjust_cell_width(sheet)

            for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.fill = self.orange_colour

            workbook.save(self.excel_file_name)
            print(f"Successfully updated excel sheet frames design for {self.excel_file_name}")

        except Exception as err:
            print(f"Exception occurred while updating excel sheet frames design for {self.excel_file_name}: {err}")

    def extracting_values_from_log_files(self):
        column_num = 3
        for html_file, column_name in zip(self.log_files_list, self.column_name_list):
            html_file_full_path = os.path.join(self.log_directory_path,html_file)
            print(f"Reading html file is {html_file_full_path}")
            with open(html_file_full_path, 'r') as file:
                html_content = file.read()
                Test_status_list = []

                for i in range(1, 6):
                    for j in range(1, 101):
                        test_name = f'TG{i}-{j:02}'
                        for line_number, line in enumerate(html_content.splitlines(), start=1):
                            if test_name in line and ("Test Passed" in line or "Test Failed" in line or "TG3-27"):
                                print(line)
                                if test_name == "TG2-19":
                                    print("TG2-19 found")
                                    required_status = line.split(" ")[4:]
                                elif test_name == "TG3-27":
                                    if "Device Supports Hibernate Mode" in line:
                                        print("Hibernate found")
                                        status = "PASS"
                                        print(status)
                                    else:
                                        status = "FAIL"
                                        status = test_name + "" + status
                                        print(status)

                                    print(f"Test {test_name} Status: {status}")
                                    Test_status_list.append(status)
                                    break
                                else:
                                    required_status = line.split(" ")[5:]
                                status = required_status[-1].strip('",')
                                status = status.replace("Passed", "PASS").replace("Failed", "FAIL")
                                print(f"Test {test_name} Status: {status}")
                                if status[4] == '"':
                                    status = test_name + "" + status.split('"')[0]
                                    Test_status_list.append(status)
                                elif status[4] == '<':
                                    status = test_name + "" + status.split('<')[0]
                                    Test_status_list.append(status)
                                break
                            else:
                                pass
                print(Test_status_list)
                print(len(Test_status_list))
                try:
                    wb = openpyxl.load_workbook(self.excel_file_name)
                    sheet = wb.active
                    for idx, value in enumerate(Test_status_list, start=2):
                        if value[-4:] == "FAIL":
                            cell = sheet.cell(row=idx, column=column_num, value=value)
                            print("Failed status encounter")
                            cell.font = openpyxl.styles.Font(color="FF0000", bold=True)
                        else:
                            sheet.cell(row=idx, column=column_num, value=value[-4:])

                    wb.save(self.excel_file_name)
                    print(f"Values appended to {self.excel_file_name}")
                except Exception as err:
                    print(f"{err}")

            column_num += 1
            print(column_num)

    def extracting_test_name_from_log_files(self):
        column_num  = 3
        for html_file, column_name in zip(self.log_files_list, self.column_name_list):
            html_file_full_path = os.path.join(self.log_directory_path, html_file)
            print(f"Reading html file is {html_file_full_path}")

        with open(html_file_full_path, 'r') as file:
            html_content = file.read()
            Test_Name_list = []
            for i in range(1, 6):
                for j in range(1, 101):
                    test_name = f'TG {i}-{j:02}'
                    for line_number, line in enumerate(html_content.splitlines(), start=1):
                        if test_name in line:
                            print(line)
                            if "***Executing" not in line:
                                if test_name == "TG 2-30" or test_name == "TG 2-31":
                                    print("TG 2-30 and TG 2-31 found")
                                    start_index = line.find("- PHY")
                                    end_index = line.find("</span>", start_index)
                                    found_string = f"{test_name} {line[start_index:end_index].strip()}"
                                    Test_Name_list.append(found_string)
                                else:
                                    # Extract the text after '-->' and before '</span>'
                                    start_index = line.find("-->") + 3
                                    end_index = line.find("</span>", start_index)
                                    found_string = f"{test_name} {line[start_index:end_index].strip()}"
                                    Test_Name_list.append(found_string)

                            else:
                                # Extract the text after 'TG' until the next HTML tag starts
                                start_index = line.find(test_name) + len(test_name)
                                end_index = line.find("<", start_index)
                                found_string = f"{test_name} {line[start_index:end_index].strip()}"
                                Test_Name_list.append(found_string)
                            break
            print(Test_Name_list)
            print(len(Test_Name_list))
            try:
                wb = openpyxl.load_workbook(self.excel_file_name)
                sheet = wb.active
                column_name = "Test Name"
                header_values = [cell.value for cell in sheet[1]]
                column_index_test_name = header_values.index(column_name) + 1

                # Append Test Name values to the "Test Name" column
                for idx, value in enumerate(Test_Name_list, start=2):
                    sheet.cell(row=idx, column=column_index_test_name, value=value)

                wb.save(self.excel_file_name)
                print(f"Values appended to {self.excel_file_name}")
            except Exception as err:
                print(f"{err}")

    def creating_sample_column(self):
        try:
            column_num = 3
            wb = openpyxl.load_workbook(self.excel_file_name)
            sheet = wb.active
            for col,log_name in zip((range(column_num, column_num + self.log_list_len)), self.log_files_list):
                current_col = openpyxl.utils.get_column_letter(col)
                log_name = int(log_name.split(".")[0])
                column_name = f"Status(Sample-{log_name})"
                self.column_name_list.append(column_name)
                sheet[f"{current_col}1"].value = column_name
                sheet[f"{current_col}1"].font = Font(bold =True)

            wb.save(self.excel_file_name)
            print("Successfully update the column name as log file names")
        except Exception as err:
            print(f"Error occurred while creating sample column is {err}")

    def iterating_over_log_directory(self):
        try:
            self.log_files_list = os.listdir(self.log_directory_path)
            self.log_files_list = [file for file in self.log_files_list if file.endswith('.htm')]
            self.log_files_list = sorted(self.log_files_list, key = lambda x: int(x.split(".")[0]))
            self.log_list_len = len(self.log_files_list)
            print(f"List of Log files are {self.log_files_list}")
            print(f"Total no. of log files are  {self.log_list_len}")
        except FileNotFoundError:
            print("File are not found or provided path is incorrect")
        except Exception as err:
            print(f"Error occurred while iteration log files is {err}")


    def creating_UHSIICompliance_excel_template(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            sheet['A1'] = "Index"
            sheet['A1'].font = Font(bold = True)
            sheet['B1'] = "Test Name"
            sheet['B1'].font = Font(bold =True)

            workbook.save(self.excel_file_name)
            print("Successfully created the UHSII Compliance Excel Template")

        except Exception as err:
            print(f"Error observed while creating the {self.excel_file_name} file")


def input_from_cmd_line():
    print("=" * 55 + "UHS II Compliance Parsing Started" + "=" * 55)
    print("file name for generating UHSII Compliance file is : UHSII_Compliance")
    excel_file_name = input("Enter the excel file name: ")
    log_directory_path = input("Enter the path of log files: ")
    return excel_file_name, log_directory_path

if __name__ == "__main__":
    excel_file_name, log_directory_path = input_from_cmd_line()
    print(f"Entered Speed Grade is {excel_file_name}")
    print(f"Entered log files directory is {log_directory_path}")
    UHSII_C = UHSIICompliance(excel_file_name, log_directory_path)
    UHSII_C.creating_UHSIICompliance_excel_template()
    UHSII_C.iterating_over_log_directory()
    UHSII_C.creating_sample_column()
    UHSII_C.extracting_test_name_from_log_files()
    UHSII_C.extracting_values_from_log_files()
    UHSII_C.excel_styling()

