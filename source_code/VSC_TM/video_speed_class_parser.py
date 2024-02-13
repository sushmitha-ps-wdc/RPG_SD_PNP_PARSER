"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:21-11-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import numpy as np
import openpyxl
import os
from openpyxl.styles import Font, PatternFill
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment, Border, Side

class VideoSpeedClassParser:

    def __init__(self, template, path_for_log):
        self.parser_template = template
        self.log_directory_path = path_for_log
        self.excel_file_path = os.path.join(self.log_directory_path, self.parser_template + '.xlsx')
        self.excel_filename = str()
        self.vsc_logs_list =[]
        self.log_list = []
        self.log_list_length = int()
        self.excel_file_path = os.path.join(self.log_directory_path, self.parser_template + '.xlsx')
        self.start_row_index = 2
        self.string_to_find = ["*** Pw(card)", "*** Pr(card)"]
        self.orange_colour = PatternFill(start_color="FFD27F", end_color="FFD27F", fill_type="solid")
        self.red_colour = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self.red_font = Font(color='00FF0000', bold=True)
        self.green_font = Font(color='006400', bold=True)


    @staticmethod
    def extract_values(soup, string_to_find):
        """
        Function to extract the values from html file by search "String_to_find" and split by "=" operator
        """
        values = []
        for tag in soup.find_all(string=lambda x: x and string_to_find in x):
            value = tag.split('=')[-1].strip()
            values.append(value)
        return values

    def excel_styling(self):
        """
        Definition for updating excel sheet with merging cells,  font bold, colouring, center alignmnet of text
        """

        def fill_first_row_with_color(sheet, color):
            for cell in sheet[1]:
                if cell.value is not None:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        def adjust_cell_width(sheet):
            for col in sheet.columns:
                max_length = 0
                column = openpyxl.utils.get_column_letter(col[0].column)
                for cells in col:
                    try:
                        if len(str(cells.value)) > max_length:
                            max_length = len(str(cells.value))
                    except Exception as e:
                        print(f"{e}")
                        pass
                adjusted_width = (max_length + 2) * 0.9
                sheet.column_dimensions[column].width = adjusted_width

        def apply_thin_borders(sheet):
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_centre_alignment(sheet):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        def bold_font_cells(sheet, bold_cells):
            for cells in bold_cells:
                sheet[cells].font = openpyxl.styles.Font(bold=True)
        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)
            sheet = workbook.active
            adjust_cell_width(sheet)
            apply_thin_borders(sheet)
            apply_centre_alignment(sheet)
            bold_font_cell_list = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']
            bold_font_cells(sheet, bold_font_cell_list)
            fill_first_row_with_color(sheet, "FFD27F")

            def merge_column_A(start_row, end_row, rows_per_merge):
                for i in range(start_row, end_row + 1, rows_per_merge):
                    merge_range = f"A{i}:A{i + rows_per_merge - 1}"
                    sheet.merge_cells(merge_range)
            def merge_column_B_and_D(start_row, end_row, rows_per_merge):
                for i in range(start_row, end_row + 1, rows_per_merge):
                    merge_range_B = f"B{i}:B{i + rows_per_merge - 1}"
                    merge_range_D = f"D{i}:D{i + rows_per_merge - 1}"
                    sheet.merge_cells(merge_range_B)
                    sheet.merge_cells(merge_range_D)

            if self.parser_template == "VSC_10":
                start_row = 2
                end_row = sheet.max_row
                rows_per_merge_column_A = 4
                rows_per_merge_column_B_and_D = 2

                merge_column_A(start_row,end_row,rows_per_merge_column_A)
                merge_column_B_and_D(start_row,end_row,rows_per_merge_column_B_and_D)

            elif self.parser_template == "VSC_30":
                start_row = 2
                end_row = sheet.max_row
                rows_per_merge_column_A = 6
                rows_per_merge_column_B_and_D = 2

                merge_column_A(start_row, end_row, rows_per_merge_column_A)
                merge_column_B_and_D(start_row, end_row, rows_per_merge_column_B_and_D)

            workbook.save(self.excel_file_path)
            print(f"Specified Values updated in {self.parser_template} excel file")

        except FileNotFoundError:
            print("File not found or paths are incorrect")

    def validating_with_specified_values(self):
        try:
            wb = openpyxl.load_workbook(self.excel_file_path)
            sheet = wb.active

            def update_result(specified_value):
                if sheet[f"E{row_num}"].value is not None:
                    if sheet[f"E{row_num}"].value >= specified_value:
                        sheet[f"F{row_num}"] = "PASS"
                    else:
                        sheet[f"F{row_num}"] = "FAIL"
                        sheet[f"F{row_num}"].font = self.red_font
                        sheet[f"E{row_num}"].font = self.red_font
                else:
                    sheet[f"F{row_num}"] = "FAIL"
                    sheet[f"F{row_num}"].font = self.red_font
                    sheet[f"E{row_num}"].fill = self.red_colour


            if self.parser_template == "VSC_10":
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2),
                                              start=2):
                    for cell in row:
                        if cell.value == "VSC6":
                            update_result(6)
                        elif cell.value == "VSC10":
                            update_result(10)
                        elif cell.value is None:
                            if sheet[f"E{row_num}"].value is not None:
                                c_index = sheet[f"D{row_num-1}"].value
                                c_index = float(c_index.split("[")[0][1:])

                                if sheet[f"E{row_num}"].value >= c_index:
                                    sheet[f"F{row_num}"] = "PASS"
                                else:
                                    sheet[f"F{row_num}"] = "FAIL"
                                    sheet[f"F{row_num}"].font = self.red_font
                                    sheet[f"E{row_num}"].font = self.red_font
                            else:
                                sheet[f"F{row_num}"] = "FAIL"
                                sheet[f"F{row_num}"].font = self.red_font
                                sheet[f"E{row_num}"].fill = self.red_colour

            elif self.parser_template == "VSC_30":
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2),
                                              start=2):
                    for cell in row:
                        if cell.value == "VSC6":
                            update_result(6)
                        elif cell.value == "VSC10":
                            update_result(10)
                        elif cell.value == "VSC30":
                            update_result(30)
                        elif cell.value is None:
                            if sheet[f"E{row_num}"].value is not None:
                                c_index = sheet[f"D{row_num - 1}"].value
                                c_index = float(c_index.split("[")[0][1:])
                                if sheet[f"E{row_num}"].value >= c_index:
                                    sheet[f"F{row_num}"] = "PASS"
                                else:
                                    sheet[f"F{row_num}"] = "FAIL"
                                    sheet[f"F{row_num}"].font = self.red_font
                                    sheet[f"E{row_num}"].font = self.red_font
                            else:
                                sheet[f"F{row_num}"] = "FAIL"
                                sheet[f"F{row_num}"].font = self.red_font
                                sheet[f"E{row_num}"].fill = self.red_colour

            wb.save(self.excel_file_path)
            print(f"Specified Values updated in {self.parser_template} excel file")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred during updating Specified Values column: {e}")

    def extracting_wr_values_from_log_file(self):
        column_num = 4
        row_indices = {}
        if self.parser_template == "VSC_10":
            row_indices = {
                '*** Pw(card)': [0, 2],
                '*** Pr(card)': [1, 3]
            }
        elif self.parser_template == "VSC_30":
            row_indices = {
                '*** Pw(card)': [0, 2, 4],
                '*** Pr(card)': [1, 3, 5]
            }
        for html_file_name in self.log_list:
            print(f"current html file {html_file_name}")
            full_log_file_path = os.path.join(self.log_directory_path, html_file_name)
            print(f"complete html file path {full_log_file_path}")

            with open(full_log_file_path, 'r') as file:
                html_content = file.read()
                values_dict = {"*** Pw(card)": [], "*** Pr(card)": []}
                soup = BeautifulSoup(html_content, 'html.parser')
                for string_to_find in self.string_to_find:
                    values_dict[string_to_find] = self.extract_values(soup, string_to_find)
                try:
                    df = pd.read_excel(self.excel_file_path)
                    for string, indices in row_indices.items():
                        values = values_dict[string]
                        print(f"Extracted values for {string}:", values)

                        for i, value in zip(indices, values):
                            if value == "":
                                df.iloc[i, column_num] = np.NAN
                            if value is not None and value != "":
                                numeric_value = round(float(str(value).split("(")[0]), 2)
                                df.iloc[i, column_num] = numeric_value
                        df.iloc[:, column_num] = pd.to_numeric(df.iloc[:, column_num], errors='coerce')

                except Exception as e:
                    print("Error handling the Excel file:", e)
                print("DataFrame shape:", df.shape)
                df.to_excel(self.excel_file_path, index=False)
                print("Excel file updated successfully.")

            if self.parser_template == "VSC_10":
                row_indices = {key: [i + 4 for i in value] for key, value in row_indices.items()}
            elif self.parser_template == "VSC_30":
                row_indices = {key: [i + 6 for i in value] for key, value in row_indices.items()}

    def update_specified_values(self):
        try:
            wb = openpyxl.load_workbook(self.excel_file_path)
            sheet = wb.active
            if self.parser_template == "VSC_10":
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2),
                                              start=2):
                    for cell in row:
                        if cell.value == "VSC6":
                            sheet[f"D{row_num}"] = "≥6[MB/s]"
                        elif cell.value == "VSC10":
                            sheet[f"D{row_num}"] = "≥10[MB/s]"
                        else:
                            sheet[f"D{row_num}"] = " "
            elif self.parser_template == "VSC_30":
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2),
                                              start=2):
                    for cell in row:
                        if cell.value == "VSC6":
                            sheet[f"D{row_num}"] = "≥6[MB/s]"
                        elif cell.value == "VSC10":
                            sheet[f"D{row_num}"] = "≥10[MB/s]"
                        elif cell.value == "VSC30":
                            sheet[f"D{row_num}"] = "≥30[MB/s]"
                        else:
                            sheet[f"D{row_num}"] = " "
            wb.save(self.excel_file_path)
            print(f"Specified Values updated in {self.parser_template} excel file")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred during updating Specified Values column: {e}")

    def update_write_read_string(self):
        try:
            wb = openpyxl.load_workbook(self.excel_file_path)
            sheet = wb.active
            if self.parser_template == "VSC_10":
                for value_row_number in range(2, sheet.max_row + 2, 1):
                    if value_row_number % 2 == 0:
                        sheet["C" + str(value_row_number)].value = "Write"
                    else:
                        sheet["C" + str(value_row_number)].value = "Read"
                print("write and read strings have been updated successfully")
            elif self.parser_template == "VSC_30":
                for value_row_number in range(2, sheet.max_row + 2, 1):
                    if value_row_number % 2 == 0:
                        sheet["C" + str(value_row_number)].value = "Write"
                    else:
                        sheet["C" + str(value_row_number)].value = "Read"
            wb.save(self.excel_file_path)
            print(f"Write and read text is updated in {self.parser_template} excel file")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred during updating Write read text column: {e}")

    def update_vsc_type(self):
        try:
            wb = openpyxl.load_workbook(self.excel_file_path)
            sheet = wb.active
            if self.parser_template == "VSC_10":
                for type_row_number in range(2, sheet.max_row + 4, 2):
                    if type_row_number % 4 == 0:
                        sheet["B" + str(type_row_number)].value = "VSC10"
                    else:
                        sheet["B" + str(type_row_number)].value = "VSC6"
                print("Type of VSC have been updated successfully")

            elif self.parser_template == "VSC_30":
                type_values_to_set = ["VSC6", "VSC10", "VSC30"]
                for index in range(2, sheet.max_row + 6, 2):
                    value_index = (index // 2 - 1) % len(type_values_to_set)
                    sheet["B" + str(index)].value = type_values_to_set[value_index]

            wb.save(self.excel_file_path)
            print(f"VCS Type is updated in {self.parser_template} excel file")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred during updating VSC Type column: {e}")

    def update_sample_num_column(self):
        try:
            wb = openpyxl.load_workbook(self.excel_file_path)
            sheet = wb.active
            if self.parser_template == "VSC_10":
                for log_file_name in self.log_list:
                    row_name = int(log_file_name.split(".")[0])
                    sheet["A" + f"{self.start_row_index}"].value = row_name
                    sheet["A" + f"{self.start_row_index}"].font = Font(bold=True)
                    self.start_row_index += 4
                print("Sample No. have been updated successfully")
            elif self.parser_template == "VSC_30":
                for log_file_name in self.log_list:
                    row_name = int(log_file_name.split(".")[0])
                    sheet["A" + f"{self.start_row_index}"].value = row_name
                    sheet["A" + f"{self.start_row_index}"].font = Font(bold=True)
                    self.start_row_index += 6
            wb.save(self.excel_file_path)
            print(f"{self.log_list_length} samples created from row '{self.start_row_index}' in the Excel file.")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred during updating sample number column: {e}")

    def iterating_over_log_directory(self):
        """
        Defintion to get the list of log_files_names and total no. of log file
        """
        try:
            self.vsc_logs_list = os.listdir(self.log_directory_path )
            self.log_list = [file for file in self.vsc_logs_list if file.endswith('htm')]
            self.log_list = sorted(self.log_list, key=lambda x: int(x.split('.')[0]))
            self.log_list_length = len(self.log_list)
            print(f"Total number of log files in a directory: {self.log_list_length}")
            print(f"List of log files are {self.log_list}")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as err:
            print(f"Occurred exception while iterating log directory is {err}")

    def create_specified_excel_structure(self):
        """
        Definition to create basic specified excel sheets header
        """
        if self.parser_template == "VSC_10":
            self.excel_filename = "VSC_10.xlsx"
        elif self.parser_template == "VSC_30":
            self.excel_filename = "VSC_30.xlsx"
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet['A1'] = "Sample No."
            sheet['B1'] = "Type"
            sheet['C1'] = "Value"
            sheet['D1'] = "Specification in [MB/s]"
            sheet['E1'] = "Actual Values in [MB/s]"
            sheet['F1'] = "Result"
            workbook.save(self.excel_file_path)
            print("successfully created header for Sequential Excel file")
        except Exception as err:
            print(f"Exception occurred at creating header for Sequential Excel file {err}")

def inputs_from_cmd_line():
    print("=" * 55 + "Command line inputs for VSC Parsing:" + "=" * 55 + "\n")
    print("Choose the below excel sheets to generate:")
    print("VSC_10\n VSC_30")
    parsing_template_input = input("Enter the required parsing file: ")
    path_for_log_files_input = input("Enter the log files directory path: ")
    print("Capturing inputs for VSC parsing are done. Below are the inputs for parsing")
    return parsing_template_input, path_for_log_files_input

if __name__ == "__main__":
    print("=" * 55 + "Started VSC Parsing" + "=" * 55 + "\n")
    parsing_template, path_for_log_files = inputs_from_cmd_line()
    print("\n\r")
    print(f"Entered parsing template is: {parsing_template}")
    print(f"Entered log files directory path is: {path_for_log_files}")
    print("\n\r")
    vsc = VideoSpeedClassParser(parsing_template,path_for_log_files)
    vsc.create_specified_excel_structure()
    vsc.iterating_over_log_directory()
    vsc.update_sample_num_column()
    vsc.update_vsc_type()
    vsc.update_write_read_string()
    vsc.update_specified_values()
    vsc.extracting_wr_values_from_log_file()
    vsc.validating_with_specified_values()
    vsc.excel_styling()
