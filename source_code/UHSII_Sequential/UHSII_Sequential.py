"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:18-12-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter


class UHSIISequential:

    def __init__(self,excel_file_name, log_directory_path):
        self.file_name = excel_file_name
        self.log_directory_path = log_directory_path
        self.excel_file_name = os.path.join(self.log_directory_path, self.file_name + ".xlsx")
        self.working_directory = os.getcwd()
        self.initial_excel_file = os.path.join(self.working_directory, self.file_name + ".xlsx")
        self.log_files_list = []
        self.column_name_list = []
        self.log_list_len = int()
        self.log_files = str()
        self.html_file_path = str()
        self.write_string_to_find = ["FM_FD WRITE Throughput [MB/s] - (1KB=1000B) : ",
                                "FM_HD WRITE Throughput [MB/s] - (1KB=1000B) : ",
                                "LPM_FD WRITE Throughput [MB/s] - (1KB=1000B) : ",
                                "LPM_HD WRITE Throughput [MB/s] - (1KB=1000B) : "]
        self.read_string_to_find = ["FM_FD READ Throughput [MB/s] - (1KB=1000B) : ",
                                "FM_HD READ Throughput [MB/s] - (1KB=1000B) : ",
                                "LPM_FD READ Throughput [MB/s] - (1KB=1000B) : ",
                                "LPM_HD READ Throughput [MB/s] - (1KB=1000B) : "]
        self.start_column_index = 7
        self.pw_performance_values = []
        self.pr_performance_values = []
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self.blue_colour = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        self.yellow_colour = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    @staticmethod
    def extract_values(soup, string_to_find):
        performance_values = []
        for label_tag in soup.find_all('span', class_='us', string=string_to_find):
            div_tag = label_tag.find_parent('div')
            if div_tag:
                next_div_tag = div_tag.find_next('div')
                if next_div_tag:
                    value_tag = next_div_tag.find('span', class_='us')
                    if value_tag:
                        int_value = value_tag.text.strip()
                        if int_value != "":
                            try:
                                int_value = round(float(int_value), 2)
                                #print(int_value)
                                performance_values.append(int_value)
                            except ValueError:
                                print(f"Error: Unable to convert '{int_value}' to float.")
                                performance_values.append(float(0))
                        else:
                            performance_values.append(" ")
                    else:
                        performance_values.append(" ")
                        print("No values found")
        return performance_values

    def excel_styling(self):

        def adjust_cell_width():
            start_column = 7
            end_column = sheet.max_column
            columns_list = list(sheet.columns)
            for col in columns_list[start_column - 1:end_column]:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                        print(f"{e}")
                        pass
                adjusted_width = (max_length + 2) * 0.9
                sheet.column_dimensions[cell.column_letter].width = adjusted_width

        def apply_thin_borders():
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=7, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_centre_alignment():
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=7, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        def merge_consecutive_cells():
            start_col =7
            max_col = sheet.max_column

            while start_col < max_col:
                end_col = start_col + 1
                sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                start_col += 2
        try:
            workbook = openpyxl.load_workbook(self.excel_file_name)
            sheet = workbook.active
            adjust_cell_width()
            apply_thin_borders()
            apply_centre_alignment()
            merge_consecutive_cells()

            workbook.save(self.excel_file_name)
        except Exception as err:
            print(f"Exception occurred while excel styling is {err}")

    def update_values_to_excel_file(self):
        workbook = openpyxl.load_workbook(self.excel_file_name)
        sheet = workbook.active
        start_row_index = 3
        # Loop for write values
        for index, value in enumerate(self.pw_performance_values):
            cell = sheet.cell(row=start_row_index + index, column=self.start_column_index, value=value)
            if value is None or value == " ":
                cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        # Loop for read values
        start_row_index = 3
        for index, value in enumerate(self.pr_performance_values):
            cell = sheet.cell(row=start_row_index + index, column=self.start_column_index + 1, value=value)
            if value is None or value == " ":
                cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        self.start_column_index += 2
        workbook.save(self.excel_file_name)

    def extracting_write_read_values_from_log_file(self):
        try:
            self.log_files = os.listdir(self.log_directory_path)
            for file in self.log_files:
                if file.endswith('.htm'):
                    self.log_files = file
                    self.html_file_path = os.path.join(self.log_directory_path, self.log_files)
                    print(self.html_file_path)
                    with open(self.html_file_path, 'r') as f:
                        html_content = f.read()
                        soup = BeautifulSoup(html_content, 'html.parser')
                        self.pw_performance_values = self.extract_values(soup, self.write_string_to_find)
                        print(f"Write Throughput values are: {self.pw_performance_values}")
                        print(f"Total No. of Write Throughput values are: {len(self.pw_performance_values)}")
                        self.pr_performance_values = self.extract_values(soup, self.read_string_to_find)
                        print(f"Read Throughput values are: {self.pr_performance_values}")
                        print(f"Total No. of Read Throughput values are: {len(self.pr_performance_values)}")
                        self.update_values_to_excel_file()
        except Exception as err:
            print(f"Exception occurred while extracting write and read values is {err}")

    def creating_sample_column(self):
        try:
            column_num = 7
            wb = openpyxl.load_workbook(self.excel_file_name)
            sheet = wb.active
            for col, log_name in zip(range(column_num, column_num + 2 * self.log_list_len, 2), self.log_files_list):
                current_col = openpyxl.utils.get_column_letter(col)
                log_name = int(log_name.split(".")[0])
                column_name = f"Sample-{log_name}"
                self.column_name_list.append(column_name)
                sheet[f"{current_col}1"].value = column_name
                sheet[f"{current_col}1"].font = Font(bold =True)
                sheet[f"{current_col}1"].fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")

                write_string = f"Write [MB/s]"
                sheet[f"{current_col}2"].value = write_string
                sheet[f"{current_col}2"].font = Font(bold=True)
                sheet[f"{current_col}2"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            read_column = 8
            for col, log_name in zip(range(read_column, read_column + 2 * self.log_list_len, 2), self.log_files_list):
                current_col = openpyxl.utils.get_column_letter(col)
                read_string = f"Read [MB/s]"
                sheet[f"{current_col}2"].value = read_string
                sheet[f"{current_col}2"].font = Font(bold=True)
                sheet[f"{current_col}2"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            wb.save(self.excel_file_name)
            print("Successfully update the column name as log file names")
        except Exception as err:
            print(f"Error occurred while creating sample column is {err}")

    def creating_template_for_excel_sheet(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            sheet['A1'] = "Performance_Min_to_Max_Range_up to NFCU64_Test_FM-LPM"
            sheet['A1'].font = Font(bold=True)
            sheet['A1'].fill = self.blue_colour
            sheet.merge_cells("A1:F1")

            sheet['A2'] = "Type"
            sheet['B2'] = "Speed Range"
            sheet['C2'] = "Mode"
            sheet['D2'] = "NFCU"
            sheet['E2'] = "TLEN"
            sheet['F2'] = "Data Size (MB)"

            header_cells = ['A2', 'B2', 'C2', 'D2', 'E2', 'F2']
            for cell in header_cells:
                sheet[cell].font = Font(bold=True)
                sheet[cell].fill = self.yellow_colour

            values_A = ["Full Duplex"] * 12 + ["Half Duplex"] * 12 + ["Full Duplex"] * 6 + ["Half Duplex"] * 6
            for i, value in enumerate(values_A, start=3):
                sheet['A' + str(i)] = value

            values_B = ["Fast Mode"] * 24 + ["Low Power Mode"] * 12
            for i, value in enumerate(values_B, start=3):
                sheet['B' + str(i)] = value

            values_C = ["Range A"] * 3 + ["Range B"] * 3
            for i in range(3, 39):
                sheet['C' + str(i)] = values_C[i % 6]

            values_D = [32, 32, 32, 32, 32, 32, 64, 64, 64, 64, 64, 64, 32, 32, 32, 32, 32, 32, 64, 64, 64, 64, 64, 64,
                        32, 32, 32, 64, 64, 64, 32, 32, 32, 64, 64, 64]
            for i, value in enumerate(values_D, start=3):
                sheet['D' + str(i)] = value

            values_E = [32768, 65536, 131072] * 12
            for i, value in enumerate(values_E, start=3):
                sheet['E' + str(i)] = value

            values_F = [16, 32, 64] * 12
            for i, value in enumerate(values_F, start=3):
                sheet['F' + str(i)] = value

            column_widths = {'A': 10, 'B': 16, 'C': 10, 'D': 5, 'E': 10, 'F': 14}
            for column, width in column_widths.items():
                sheet.column_dimensions[column].width = width

            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border

            workbook.save(self.excel_file_name)
            print("Successfully created the UHSII Compliance Excel Template")

        except Exception as err:
            print(f"Error observed while creating the {self.excel_file_name} file")


    def iterating_over_log_directory(self):
        try:
            self.log_files_list = os.listdir(self.log_directory_path)
            self.log_files_list = [file for file in self.log_files_list if file.endswith('.htm')]
            self.log_files_list = sorted(self.log_files_list, key = lambda x: int(x.split(".")[0]))
            self.log_list_len = len(self.log_files_list)
            print(f"List of Log files are {self.log_files_list}")
            print(f"Total no. of log files are  {self.log_list_len}")
        except FileNotFoundError:
            print("File are not found or provided path is incorrect")
        except Exception as err:
            print(f"Error occurred while iteration log files is {err}")


def input_from_cmd_line():
    print("=" * 55 + "UHS-II Sequential Parsing Started" + "=" * 55)
    print("File name for generating UHS-II Sequential file is : UHSII_Sequential")
    excel_file_name = input("Enter the excel file name: ")
    log_directory_path = input("Enter the path of log files: ")
    return excel_file_name, log_directory_path

if __name__ == "__main__":
    excel_file_name, log_directory_path = input_from_cmd_line()
    print(f"Entered VSC Full Capacity is {excel_file_name}")
    print(f"Entered log files directory is {log_directory_path}")
    UHSII_seq = UHSIISequential(excel_file_name, log_directory_path)
    UHSII_seq.iterating_over_log_directory()
    UHSII_seq.creating_template_for_excel_sheet()
    UHSII_seq.creating_sample_column()
    UHSII_seq.extracting_write_read_values_from_log_file()
    UHSII_seq.excel_styling()