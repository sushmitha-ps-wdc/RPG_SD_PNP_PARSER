"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:23-11-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment, Border, Side
from bs4 import BeautifulSoup
import pandas as pd
import re

class SpeedGradeParser:

    def __init__(self, speed_grade, log_directory_path):
        self.speed_grade = speed_grade
        self.log_directory_path = log_directory_path
        self.excel_file_name = os.path.join(self.log_directory_path,self.speed_grade +".xlsx")
        self.log_files_list = []
        self.log_list_len = int()
        self.start_column_index = 'D'
        self.start_column_num = 3
        self.column_name_list = []
        self.strings_to_find = ["*** Pw(card)", "*** Tfw(avg)max", "*** Tfw(max)max","*** Tfr(4KB)max","*** Pr(card)"]
        self.start_row_num = 3
        self.red_font = Font(color='00FF0000', bold=True)
        self.green_font = Font(color='006400', bold=True)
        self.light_blue_color = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        self.dark_yellow_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.grey_color = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self.light_green_color = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        self.red_colour = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))

    @staticmethod
    def extract_values(soup,string_to_find):
        values = []
        for tag in soup.find_all(string=lambda x: x and string_to_find in x):
            value = tag.split("=")[-1].strip()
            values.append(value)
        return values

    def excel_styling(self):

        def fill_first_row_with_color():
            for cell in sheet[1]:
                if cell.value is not None:
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type='solid')

        def adjust_cell_width():
            for col in sheet.columns:
                max_length = 0
                column = openpyxl.utils.get_column_letter(col[0].column)
                for cells in col:
                    try:
                        if len(str(cells.value)) > max_length:
                            max_length = len(str(cells.value))
                    except Exception as e:
                        print(f"{e}")
                        pass
                adjusted_width = (max_length + 2) * 0.9
                sheet.column_dimensions[column].width = adjusted_width

        def apply_thin_borders():
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_centre_alignment():
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        try:
            workbook = openpyxl.load_workbook(self.excel_file_name)
            sheet = workbook.active
            fill_first_row_with_color()
            adjust_cell_width()
            apply_thin_borders()
            apply_centre_alignment()
            sheet.merge_cells("A1:A14")

            max_column = sheet.max_column

            def yellow_colouring_for_merged_cells(yellow_row_num):
                """
                Definition to fill yellow colour for merged rows 2,7 and 12
                """
                for cell in sheet[yellow_row_num][1:max_column + 1]:
                    sheet.merge_cells(start_row=yellow_row_num, start_column=cell.column, end_row=yellow_row_num, end_column=max_column)
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.fill = yellow_fill
                    cell.font = Font(bold = True)

            yellow_colouring_for_merged_cells(2)
            yellow_colouring_for_merged_cells(7)
            yellow_colouring_for_merged_cells(12)

            def fill_color_in_range(sheet, cell_range, color):
                """
                Definition to fill a specific range of cells with a specified color
                """
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for row in sheet[cell_range]:
                    for cell in row:
                        cell.fill = fill

            light_grey_ranges = ['B3:B6', 'B8:B11', 'B13:B14']
            light_green_ranges = ['C3:C6', 'C8:C11', 'C13:C14']
            for cell_range in light_grey_ranges:
                fill_color_in_range(sheet, cell_range, "D9D9D9")  # Light Blue color
            for cell_range in light_green_ranges:
                fill_color_in_range(sheet, cell_range, "E2EFDA")  # Light green color

            workbook.save(self.excel_file_name)
        except Exception as err:
            print(f"Error occurred while updating excel sheet format is {err}")

    def validating_extracted_values_with_specified_values(self):
        column_num = 4
        try:
            wb = openpyxl.load_workbook(self.excel_file_name)
            sheet = wb.active
            for rows in sheet.iter_rows():
                for cell in rows:
                    if cell.value is not None and isinstance(cell.value, str):
                        cell.value = cell.value.replace("(", "[").replace(")", "]")
                        cell.value = cell.value.replace("Âµs", "µs")
            wb.save(self.excel_file_name)

            df = pd.read_excel(self.excel_file_name, na_filter=False, engine="openpyxl")
            specified_values_into_list = df['Specified'].to_list()
            print(f"Specified values are {specified_values_into_list}")

            wb = openpyxl.load_workbook(self.excel_file_name)
            sheet = wb.active
            for column_name in self.column_name_list:
                print(f"Validating the Sample values: {column_name}")
                values_into_list = df[column_name].tolist()
                for i, (actual_value, spec_value) in enumerate(zip(values_into_list, specified_values_into_list)):
                    if i >= len(specified_values_into_list):
                        print("Index out of range for specified_values_into_list. Skipping validation.")
                        break
                    if actual_value == spec_value == '':
                        pass
                    elif actual_value == "nan" and spec_value == '':
                        pass
                    elif (actual_value == " " or actual_value is None) and spec_value is not None:
                        print(f"No values encountered.Please check Log {column_name} files")
                        sheet.cell(row=i + 2, column=column_num).fill = self.red_colour
                    else:
                        try:
                            if actual_value == '':
                                sheet.cell(row=i + 2, column=column_num).fill = self.red_colour
                            else:
                                number_from_validating_list = float(actual_value.split('[')[0])
                                unit_from_validating_list = actual_value.split('[')[1].split(']')[0]

                                if unit_from_validating_list == 'µs' or unit_from_validating_list == "us":
                                    number_from_validating_list /= 1000

                                comparison_operator = spec_value[0:1]
                                numeric_value = re.search(r'\d+(\.\d+)?', spec_value)
                                number_from_specific_list = float(numeric_value.group())

                                if comparison_operator == '≥' and number_from_validating_list >= number_from_specific_list:
                                    pass
                                elif comparison_operator == '≤' and number_from_validating_list <= number_from_specific_list:
                                    pass
                                else:
                                    print(f"Actual value is not expected to Specified Value in log file {column_name}")
                                    sheet.cell(row=i+2, column = column_num).font = self.red_font
                        except Exception as err:
                            print(f"{err}")
                column_num = column_num + 1
            wb.save(self.excel_file_name)
        except Exception as err:
            print(f"{err}")

    def extracting_values_from_log_files(self):
        column_num  = 3
        for html_file, column_name in zip(self.log_files_list, self.column_name_list):
            html_file_full_path = os.path.join(self.log_directory_path,html_file)
            print(f"Reading html file is {html_file_full_path}")
            print(f"Updating current sample values in {column_name}")
            with open(html_file_full_path, 'r', encoding="utf8") as file:
                html_content = file.read()
            values_dict = {
                "*** Pw(card)": [],
                "*** Tfw(avg)max": [],
                "*** Tfw(max)max": [],
                "*** Tfr(4KB)max": [],
                "*** Pr(card)": []
            }
            row_indices = {
                "*** Pw(card)": [1, 6],
                "*** Tfw(avg)max": [2, 7],
                "*** Tfw(max)max": [3, 8],
                "*** Tfr(4KB)max": [4, 9, 12],
                "*** Pr(card)": [11]
            }
            soup = BeautifulSoup(html_content,"html.parser")
            for string_to_find in self.strings_to_find:
                values_dict[string_to_find] = self.extract_values(soup,string_to_find)
            for string, value in values_dict.items():
                print(f"Extracted values for {string}:", value)
            try:
                df = pd.read_excel(self.excel_file_name)
                print(df.shape)
                for string, indices in row_indices.items():
                    values = values_dict[string]
                    df.iloc[:,column_num] = df.iloc[:,column_num].astype(str)
                    for i, val in zip(indices,values):
                        df.iloc[i,column_num] = str(val)
                df.to_excel(self.excel_file_name, index=False)
                print("Successfully updated the extracted values to Speed Grade File")
            except Exception as err:
                print(f"Error occurred while updating values is {err}")
            column_num += 1

    def creating_sample_column(self):
        try:
            column_num = 4
            wb = openpyxl.load_workbook(self.excel_file_name)
            sheet = wb.active
            for col,log_name in zip((range(column_num, column_num + self.log_list_len)), self.log_files_list):
                current_col = openpyxl.utils.get_column_letter(col)
                log_name = (log_name.split(".")[0])
                column_name = f"Sample-{log_name}"
                self.column_name_list.append(column_name)
                sheet[f"{current_col}1"].value = column_name
                sheet[f"{current_col}1"].font = Font(bold =True)

            wb.save(self.excel_file_name)
            print("Successfully update the column name as log file names")
        except Exception as err:
            print(f"Error occurred while creating sample column is {err}")


    def iterating_over_log_directory(self):
        try:
            self.log_files_list = os.listdir(self.log_directory_path)
            self.log_files_list = [file for file in self.log_files_list if file.endswith('.htm')]
            self.log_files_list = sorted(self.log_files_list, key = lambda x: (x.split(".")[0]))
            self.log_list_len = len(self.log_files_list)
            print(f"List of Log files are {self.log_files_list}")
            print(f"Total no. of log files are  {self.log_list_len}")
        except FileNotFoundError:
            print("File are not found or provided path is incorrect")
        except Exception as err:
            print(f"Error occured while iteration log files is {err}")


    def creating_speed_grade_excel_template(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            if self.speed_grade == "U1":
                sheet['A1'] = "Grade(U1)"
                sheet['C3'] = sheet['C8'] = sheet['C13'] = "≥10[MB/s]"
            elif self.speed_grade == "U3":
                sheet['A1'] = "Grade(U3)"
                sheet['C3'] = sheet['C8'] = sheet['C13'] = "≥30[MB/s]"
            sheet.merge_cells("A1:A14")
            sheet['B1'] = "Operation"
            sheet['B2'] = "Unit of One RU (SDR50, 80MHz)"
            sheet['B3'] = sheet['B8'] = "Pw (card)"
            sheet['B4'] = sheet["B9"] = "Tfw (avg) max"
            sheet['B5'] = sheet["B10"] = "Tfw (max) max"
            sheet['B6'] = sheet['B11'] = sheet['B14'] = "Tfr (4KB) max"
            sheet['B7'] = "Multiple random RU (SDR50, 80MHz)"
            sheet['B12'] = "Pr (SDR50, 80MHz)"
            sheet['C1'] = "Specified"
            sheet['C4'] = sheet['C9'] = "≤100[ms]"
            sheet['C5'] = sheet['C10'] = "≤750[ms]"
            sheet['C6'] = sheet['C11'] = sheet['C14'] = "≤20[ms]"
            sheet['B13'] = "Pr (card)"
            workbook.save(self.excel_file_name)
            print("Successfully created the Speed Grade Excel Template")
        except Exception as err:
            print(f"Error observed while creating the {self.excel_file_name} file")

def input_from_cmd_line():
    print("=" * 55 + "Speed Garde Parsing Started" + "=" * 55)
    print("Choose below Speed Garde to generate respective file")
    print("U1\nU3\n")
    speed_grade = input("Enter the Speed Grade: ")
    log_directory_path = input("Enter the path of log files: ")
    return speed_grade, log_directory_path

if __name__ == "__main__":
    speed_grade, log_directory_path = input_from_cmd_line()
    print(f"Entered Speed Grade is {speed_grade}")
    print(f"Entered log files directory is {log_directory_path}")
    sgp = SpeedGradeParser(speed_grade, log_directory_path)
    sgp.creating_speed_grade_excel_template()
    sgp.iterating_over_log_directory()
    sgp.creating_sample_column()
    sgp.extracting_values_from_log_files()
    sgp.validating_extracted_values_with_specified_values()
    sgp.excel_styling()











