import os
import re
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd


class ComplianceParser:
    global df
    def __init__(self, parser_file_name, log_directory_path):
        self.parser_file_name = parser_file_name
        self.log_directory_path = log_directory_path
        self.file_list = []
        self.log_files = []
        self.log_list = []
        self.column_list = []
        self.file_name = str()
        self.html_file_path = str()
        self.excel_file_name = str()
        self.start_column = "D"
        self.log_list_length = int()
        self.strings_to_find_sdxc = ["*** Pw(card)", "*** Tfw(avg)max", "*** Tfw(max)max", "*** Tfr(4KB)max",
                                     "*** Pr(card)"]
        self.strings_to_find_sdhc = ["*** Pw(card)", "*** Tfw(avg)max", "*** Tfw(max)max", "*** Tfr(4KB)max",
                                     "*** Pr(card)", "*** Pc(card)"]
        self.red_font = Font(color='00FF0000', bold =True)
        self.green_font = Font(color='006400', bold=True)
        self.redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

    @staticmethod
    def extract_values(soup, string_to_find):
        values = []
        for tag in soup.find_all(string=lambda x: x and string_to_find in x):
            value = tag.split('=')[-1].strip()
            values.append(value)
        return values

    def excel_style_update(self):
        def fill_first_row_with_color(sheet, color):
            for cell in sheet[1]:
                if cell.value is not None:  # Check if the cell has data
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        def adjust_cell_width():
            for col in sheet.columns:
                max_length = 0
                column = openpyxl.utils.get_column_letter(col[0].column)
                for cells in col:
                    try:
                        if len(str(cells.value)) > max_length:
                            max_length = len(str(cells.value))
                    except Exception as e:
                        print(f"{e}")
                        pass
                adjusted_width = (max_length + 2) * 0.9
                sheet.column_dimensions[column].width = adjusted_width

        if self.file_name == "SDHC":
            try:
                workbook = openpyxl.load_workbook(self.excel_file_path)
                sheet = workbook.active
                # To Merge the cells with respect to rows of SDHC Excel (Hard-coding the index values)
                sheet.merge_cells(start_row=2, start_column=1, end_row=10, end_column=1)
                sheet.merge_cells(start_row=11, start_column=1, end_row=19, end_column=1)
                sheet.merge_cells(start_row=20, start_column=1, end_row=25, end_column=1)

                # To apply middle alignment of data in merged cells of SDHC Excel (Hard-coding the index values)
                for row in sheet.iter_rows(min_row=2, max_row=24, min_col=1, max_col=1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # To apply thin borders to cells with data
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                # Hard-code the values with respect to data frames
                for row in sheet.iter_rows(min_row=2, max_row=24, min_col=0, max_col=13):
                    for cell in row:
                        if cell.value:
                            cell.border = thin_border
                adjust_cell_width()
                fill_first_row_with_color(sheet, "FFD27F")
                workbook.save(self.excel_file_path)
                print("Successfully updated excel sheet frames design")
            except Exception as err:
                print(f"Exception occurred while updating SDHC excel sheet frames design is : {err}")

        elif self.file_name == "SDXC":
            try:
                workbook = openpyxl.load_workbook(self.excel_file_path)
                sheet = workbook.active

                # To Merge the cells with respect to rows of SDXC Excel (Hard-coding the index values)
                sheet.merge_cells(start_row=2, start_column=1, end_row=11, end_column=1)
                sheet.merge_cells(start_row=12, start_column=1, end_row=21, end_column=1)

                # # To apply middle alignment of data in merged cells of SDXC Excel (Hard-coding the index values)
                for row in sheet.iter_rows(min_row=2, max_row=21, min_col=1, max_col=1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                for row in sheet.iter_rows(min_row=2, max_row=21, min_col=0, max_col=sheet.max_column):
                    for cell in row:
                        if cell.value:
                            cell.border = self.thin_border
                adjust_cell_width()
                fill_first_row_with_color(sheet, "FFD27F")

                workbook.save(self.excel_file_path)
                print("Successfully updated excel sheet frames design")
            except Exception as err:
                print(f"Exception occurred while updating SDXC excel sheet frames design is : {err}")


    def validating_with_specific_values(self):
        """
        function to perform Validation of Sample run values with specified values
        """
        column_num = 4
        df = pd.read_excel(self.excel_file_path, na_filter=False, engine='openpyxl')
        specific_values_into_list = df['Specified'].tolist()
        print(f"specific values of reading: {specific_values_into_list}")

        wb = openpyxl.load_workbook(self.excel_file_path)
        ws = wb['Sheet1']

        for column_name in self.column_list:
            print(f"Validating the Sample values: {column_name}")
            values_into_list = df[column_name].tolist()
            print(values_into_list)
            for i, value in enumerate(values_into_list):
                if value == '' or value == "[MB/s]" or value == "[ms]" or value == "[us]":
                    print("No values encountered.Please check Log files")
                    ws.cell(row=i + 2, column=column_num).value = None
                    ws.cell(row=i + 2, column=column_num).fill = self.redFill
                else:
                    number_from_validating_list = value.split('[')[0]
                    number_from_validating_list = float(number_from_validating_list)

                    unit_from_validating_list = value.split('[')[1].split(']')[0]
                    if unit_from_validating_list == 'us':
                        number_from_validating_list = number_from_validating_list / 1000
                        number_from_validating_list = float(number_from_validating_list)

                    comparison_operator = specific_values_into_list[i][0:1]
                    spec_value = specific_values_into_list[i]
                    numeric_value = re.search(r'\d+(\.\d+)?', spec_value)
                    number_from_specific_list = float(numeric_value.group())
                    ws.cell(row=i+2, column=3).font = self.green_font    # for specified values column
                    try:
                        # Performing validation based on the comparison operator
                        if comparison_operator == '≥' and number_from_validating_list >= number_from_specific_list:
                            pass
                        elif comparison_operator == '≤' and number_from_validating_list <= number_from_specific_list:
                            pass
                        else:
                            ws.cell(row=i + 2, column=column_num).font = self.red_font  # if fails
                    except Exception as err:
                        print(f"{err}")
            column_num = column_num + 1
        wb.save(self.excel_file_path)
        print("Validating the Performance values has been completed and updated to excel successfully")

    def updating_sdhc_parameter_value(self):
        column_num = 3
        for log_file, column_name in zip(self.log_list, self.column_list):
            html_file_path = f"{self.log_directory_path}/{log_file}"
            with open(html_file_path, 'r') as file:
                html_content = file.read()
            print(f"Updating current sample values in {column_name}")
            print(f"Reading values from {log_file}")
            sdhc_values_dict = {"*** Pw(card)": [], "*** Tfw(avg)max": [], "*** Tfw(max)max": [], "*** Tfr(4KB)max": [],
                           "*** Pr(card)": [], "*** Pc(card)": []}
            sdhc_row_indices = {
                '*** Pw(card)': [0, 9, 18],
                '*** Tfw(avg)max': [1, 10, 19],
                '*** Tfw(max)max': [2, 11, 20],
                '*** Tfr(4KB)max': [3, 5, 12, 14, 21, 23],
                '*** Pr(card)': [4, 13, 22],
                '*** Pc(card)': [6, 7, 8, 15, 16, 17]
            }
            soup = BeautifulSoup(html_content, 'html.parser')

            for string_to_find in self.strings_to_find_sdhc:
                sdhc_values_dict[string_to_find] = self.extract_values(soup, string_to_find)
            try:
                df = pd.read_excel(self.excel_file_path)
                for string, indices in sdhc_row_indices.items():
                    if string == "*** Pc(card)":
                        pc_card_list = sdhc_values_dict[string]
                        positions_to_remove = [3, 4, 5, 9, 10, 11]
                        positions_to_remove.sort(reverse=True)
                        for pos in positions_to_remove:
                            del pc_card_list[pos]
                    values = sdhc_values_dict[string]
                    print(f"Extracted values for {string}:", values)
                    df.iloc[:, column_num] = df.iloc[:, column_num].astype(str)
                    for i, value in zip(indices, values):
                        df.iloc[i, column_num] = str(value)
                print("DataFrame shape:", df.shape)
                df.to_excel(self.excel_file_path, index=False)
            except Exception as e:
                print("Error handling the Excel file:", e)
            print("Excel file updated successfully.")
            column_num = column_num + 1

        wb = openpyxl.load_workbook(self.excel_file_path)
        sheet = wb.active

        def converting_braces_in_cells():
            for rows in sheet.iter_rows():
                for cells in rows:
                    if cells.value is not None and isinstance(cells.value, str):
                        cells.value = cells.value.replace('(', '[').replace(')', ']')

        converting_braces_in_cells()
        wb.save(self.excel_file_path)

    def updating_sdxc_parameters_value(self):
        column_num = 3
        global df
        for log_file, column_name in zip(self.log_list, self.column_list):
            html_file_path = f"{self.log_directory_path}/{log_file}"
            with open(html_file_path, 'r') as file:
                html_content = file.read()
            print(f"Updating current sample values in {column_name}")
            print(f"Reading values from {log_file}")

            values_dict = {"*** Pw(card)": [], "*** Tfw(avg)max": [], "*** Tfw(max)max": [], "*** Tfr(4KB)max": [], "*** Pr(card)": []}
            row_indices = {
                '*** Pw(card)': [0, 4, 10, 14],
                '*** Tfw(avg)max': [1, 5, 11, 15],
                '*** Tfw(max)max': [2, 6, 12, 16],
                '*** Tfr(4KB)max': [3, 7, 9, 13, 17, 19],
                '*** Pr(card)': [8, 18]
            }
            soup = BeautifulSoup(html_content, 'html.parser')

            for string_to_find in self.strings_to_find_sdxc:
                values_dict[string_to_find] = self.extract_values(soup, string_to_find)
            try:
                df = pd.read_excel(self.excel_file_path)
                for string, indices in row_indices.items():
                    values = values_dict[string]
                    print(f"Extracted values for {string}:", values)
                    df.iloc[:, column_num] = df.iloc[:, column_num].astype(str)
                    for i, value in zip(indices, values):
                        df.iloc[i, column_num] = str(value)
            except Exception as e:
                print("Error handling the Excel file:", e)
            print("DataFrame shape:", df.shape)
            df.to_excel(self.excel_file_path, index=False)
            print("Excel file updated successfully.")
            column_num = column_num + 1

        wb = openpyxl.load_workbook(self.excel_file_path)
        sheet = wb.active

        def converting_braces_in_cells():
            for rows in sheet.iter_rows():
                for cells in rows:
                    if cells.value is not None and isinstance(cells.value, str):
                        cells.value = cells.value.replace('(', '[').replace(')', ']')

        converting_braces_in_cells()
        wb.save(self.excel_file_path)

    def fetching_parameter_values_from_log(self):
        if self.file_name == "SDHC":
            self.updating_sdhc_parameter_value()
        elif self.file_name == "SDXC":
            self.updating_sdxc_parameters_value()

    def create_column(self):
        try:
            self.column_list = []
            wb = openpyxl.load_workbook(self.excel_file_path)
            sheet = wb.active
            start_column_index = openpyxl.utils.column_index_from_string(self.start_column)
            for volume ,name in zip((range(start_column_index, start_column_index + self.log_list_length)),self.log_list):
                current_column = openpyxl.utils.get_column_letter(volume)
                column_name = int(name.split(".")[0])
                column_name = f"Sample-{column_name}"
                self.column_list.append(column_name)
                if not sheet[f"{current_column}1"].value:
                    sheet[f"{current_column}1"] = column_name
                    sheet[f"{current_column}1"].font = Font(bold=True)
            wb.save(self.excel_file_path)
            print(f"{self.log_list_length} samples created from column '{self.start_column}' in the Excel file.")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred: {e}")

    def create_file_path(self):
        if self.file_name:
            self.excel_file_path = os.path.join(self.log_directory_path, self.file_name + ".xlsx")
        else:
            self.excel_file_path = os.path.join(self.log_directory_path, "default.xlsx")

    def creating_basic_excel_file_template(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            if self.file_name == "SDHC":
                sheet['A1'] = "Class"
                sheet['B1'] = "Parameters"
                sheet['C1'] = "Specified"
                sheet['A2'] = "CLASS 6"
                sheet['A11'] = "CLASS6 under CLASS4 Condition"
                sheet['A20'] = "CLASS 10"
                sheet['B2'] = sheet['B11'] = sheet["B20"] = "Pw(card)"
                sheet['B3'] = sheet['B12'] = sheet["B21"] = "Tfw(avg)"
                sheet["B4"] = sheet['B13'] = sheet["B22"] = "Tfw(max)"
                sheet["B5"] = sheet["B14"] = sheet["B23"] = "Tfr(4KB)"
                sheet["B6"] = sheet["B15"] = sheet["B24"] = "Pr(card)"
                sheet["B7"] = sheet["B16"] = sheet["B25"] = "Tfr(4KB)"
                sheet["B8"] = sheet["B17"] = "Pc(card)(25%)"
                sheet["B9"] = sheet["B18"] = "Pc(card)(50%)"
                sheet["B10"] = sheet["B19"] = "Pc(card)(75%)"

                sheet['C2'] = sheet['C6'] = "≥6[MB/s]"
                sheet['C3'] = sheet['C12'] = sheet["C21"] = "≤100[ms]"
                sheet["C4"] = sheet['C13'] = sheet["C22"] = "≤750[ms]"
                sheet["C5"] = sheet["C14"] = sheet["C23"] = sheet["C7"] = sheet["C16"] = sheet["C25"] = "≤12[ms]"
                sheet["C6"] = sheet["C24"] = "≥6[MB/s]"
                sheet["C15"] = sheet['C11'] = "≥4[MB/s]"

                sheet["C8"] = "≥3.6[MB/s]"
                sheet["C9"] = "≥2[MB/s]"
                sheet["C10"] = "≥0.86[MB/s]"

                sheet["C17"] = "≥2.4[MB/s]"
                sheet["C18"] = "≥1.33[MB/s]"
                sheet["C19"] = "≥0.57[MB/s]"
                sheet["C20"] = sheet["C24"] = "≥10[MB/s]"

            elif self.file_name == "SDXC":
                sheet['A1'] = "Class"
                sheet['B1'] = "Parameters"
                sheet['C1'] = "Specified"

                sheet['A2'] = "CLASS 6"
                sheet['A12'] = "CLASS 10"

                sheet['B2'] = sheet['B12'] = "Pw(card) Unit of one RU"
                sheet['B3'] = sheet['B13'] = "Tfw(avg)max Unit of one RU"
                sheet["B4"] = sheet['B14'] = "Tfw(max)max Unit of one RU"
                sheet["B5"] = sheet["B15"] = "Tfr(4KB)max Unit of one RU"

                sheet["B6"] = sheet["B16"] = "Pw(card) Multiple of one RU"
                sheet["B7"] = sheet["B17"] = "Tfw(avg)max Multiple of one RU"
                sheet["B8"] = sheet["B18"] = "Tfw(max)max Multiple of one RU"
                sheet["B9"] = sheet["B19"] = "Tfr(4KB)max Multiple of one RU"
                sheet["B10"] = sheet["B20"] = "Pr(card)"
                sheet["B11"] = sheet["B21"] = "Tfr(4KB)max"

                sheet['C2'] = sheet['C6'] = sheet['C10'] = "≥6[MB/s]"
                sheet['C3'] = sheet['C7'] = sheet["C13"] = sheet['C17'] = "≤100[ms]"
                sheet["C4"] = sheet['C8'] = sheet["C14"] = sheet['C18'] = "≤750[ms]"
                sheet["C5"] = sheet["C9"] = sheet["C11"] = sheet["C15"] = sheet["C19"] = sheet["C21"] = "≤20[ms]"
                sheet["C12"] = sheet["C16"] = sheet["C20"] = "≥10[MB/s]"

            elif self.file_name == "SDUC":
                print("Please create a script for SDXC Compliance template")

            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.font = Font(bold=True)
            print(self.excel_file_path)
            workbook.save(self.excel_file_path)
            print("successfully created SDHC excel file")
        except Exception as err:
            print(f"Exception occurred while creating SDHC Excel file template is {err}")


    def finding_card_capacity(self):
        try:
            self.file_list = os.listdir(self.log_directory_path)
            self.log_list = [file for file in self.file_list if file.endswith('htm')]
            self.log_list = sorted(self.log_list, key=lambda x: int(x.split('.')[0]))
            self.log_list_length = len(self.log_list)
            print(f"Total number of log files in a directory: {self.log_list_length}")
            print(f"List of Log files in HTML format: {self.log_list}")
            for file in self.log_list:
                    self.html_file_path = os.path.join(self.log_directory_path, file)
                    print(f"html log file path is {self.html_file_path}")
                    with open(self.html_file_path, 'r') as f:
                        html_content = f.read()
                        pattern = r'mem_capacity\s*=\s*(\d+)\s*KBytes'
                        match = re.search(pattern, html_content)

                        if match:
                            mem_capacity_kb = int(match.group(1))
                            mem_capacity_gb = mem_capacity_kb / (1024 ** 2)
                            if mem_capacity_gb <= 32:
                                self.file_name = "SDHC"
                            elif 32 < mem_capacity_gb <= 2048:  # 2TB in GB
                                self.file_name = "SDXC"
                            elif mem_capacity_gb > 2048:
                                self.file_name = "SDUC"
                    print(f"The memory capacity is {mem_capacity_gb}GB and excel file name is {self.file_name}")

        except Exception as err:
            print(f"Exception occureed while determining mem capacity is {err}")


def process_cmd_line_inputs():
    print("=" * 55 + "Command line inputs for Compliance Parsing:" + "=" * 55 + "\n")
    print("Enter the excel file name for Compliance Parser is : Compliance ")
    parser_file_name = input("Enter the excel file name: ")
    log_directory_path = input("Enter the path of log files: ")
    return parser_file_name, log_directory_path

if __name__ == "__main__":
    parser_file_name, log_directory_path = process_cmd_line_inputs()
    print(f"Entered VSC Full Capacity is {parser_file_name}")
    print(f"Entered log files directory is {log_directory_path}")
    com_par = ComplianceParser(parser_file_name, log_directory_path)
    com_par.finding_card_capacity()
    com_par.create_file_path()
    com_par.creating_basic_excel_file_template()
    com_par.create_column()
    com_par.fetching_parameter_values_from_log()
    com_par.validating_with_specific_values()
    com_par.excel_style_update()


