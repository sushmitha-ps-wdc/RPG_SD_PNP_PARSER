"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:25-12-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border

pass_pattern = """\
   ######     ##    #####   #####  
    ##  ##   ####  ##    # ##    # 
    ##  ##  ##  ## ##      ##      
    #####   ##  ##  #####   #####  
    ##      ######      ##      ## 
    ##      ##  ## #    ## #    ## 
   ####     ##  ##  #####   #####\
"""


fail_pattern = """\
   #######    ##     ####  ####    
    ##   #   ####     ##    ##     
    ## #    ##  ##    ##    ##     
    ####    ##  ##    ##    ##     
    ## #    ######    ##    ##     
    ##      ##  ##    ##    ##  ## 
   ####     ##  ##   ####  #######\
"""

class CQParser:

    def __init__(self,excel_file_name, log_directory_path):
        self.file_name = excel_file_name
        self.log_directory_path = log_directory_path
        self.excel_file_name = os.path.join(self.log_directory_path, self.file_name + ".xlsx")
        self.log_files_list = []
        self.log_list_len = int()
        self.column_name_list = []
        self.ca_log_names = []
        self.status_list = []
        self.txt_file_full_path = str()
        self.start_index_list = []
        self.length_of_testcase_name = int()
        self.len_tc_name_list = []
        self.row_index_to_delete = []
        self.error_line = None
        self.subscripts_status_list = []
        self.delete_rows = []
        self.orange_colour = PatternFill(start_color="FFD27F", end_color="FFD27F", fill_type="solid")
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self.start_index_list_non_ca = []
        self.status_list_non_ca = []
        self.red_font = Font(bold=True, color="FF0000")
        self.saffron = Font(bold=True, color = "F79646")
        self.red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.specified_subscripts_nos = [16, 15, 10, 5, 16, 20, 16, 5, 15, 8, 18, 25, 12, 15, 10, 10, 5, 5, 13, 14,
                                         7, 6, 10, 19, 11,11,5,15,10 ] # No. of specified Subscripts for only CA based Logs.

    def excel_styling(self):
        """
        Definition to excel styling such as applying thin border, adjust cells width,
        colouring, bold font, merging cells
        """

        def adjust_cell_width(sheet):
            max_column_length = {}

            for col in sheet.columns:
                column_index = col[0].column
                max_length = max(len(str(cell.value)) for cell in col)
                max_column_length[column_index] = max_length

            for col in sheet.columns:
                column_index = col[0].column
                if column_index == 2:
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = max_column_length[
                                                                                                        column_index] * 0.9
                else:
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = max_column_length[column_index] * 1.0

        def apply_thin_borders(sheet):
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_alignment(sheet):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for col_num, cell in enumerate(row, start=1):
                    if col_num == 2 and col_num == 3:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

        def fill_first_column(sheet):
            start_row = 2
            for row_num, row in enumerate(
                    sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=1),
                    start=start_row):
                for cell in row:
                    cell.value = row_num - 1
        try:
            workbook = openpyxl.load_workbook(self.excel_file_name)
            sheet = workbook.active
            fill_first_column(sheet)
            apply_alignment(sheet)
            apply_thin_borders(sheet)
            for row_idx in self.start_index_list_non_ca:
                sheet.merge_cells(f"B{row_idx}:E{row_idx}")
            adjust_cell_width(sheet)

            for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.fill = self.orange_colour

            workbook.save(self.excel_file_name)
            print(f"Successfully updated excel sheet frames design for {self.excel_file_name}")

        except Exception as err:
            print(f"Exception occurred while updating excel sheet frames design for {self.excel_file_name}: {err}")

    def extracting_status_for_non_CA_logs(self):
        filtered_log_files = [file for file in self.log_files_list if not file.startswith('CA')]
        print(f"total no. of non CA logs are {len(filtered_log_files)}")
        for file in filtered_log_files:
            txt_file_full_path = os.path.join(self.log_directory_path, file)
            print("\n\r")
            print(f"Reading txt format log file is {file}")
            try:
                with open(txt_file_full_path, 'r') as file:
                    content = file.read()
                    self.error_line=str()
                    if pass_pattern in content:
                        self.status_list_non_ca.append("PASS")
                        print("Main log result is passed")
                    elif fail_pattern in content:
                        self.status_list_non_ca.append("FAIL")
                        print("Main log result is Failed")
                        lines = content.split('\n')
                        for i in range(len(lines) - 1):
                            if re.search("TestFailedError:", lines[i]):
                                error_line = lines[i + 1].strip()
                                print(f"Next line after 'TestFailedError': {error_line}")
                                self.error_line_list.append(error_line)
                                break
                    elif pass_pattern not in content and fail_pattern not in content:
                        self.status_list_non_ca.append("STOPPED")
                        print("Main log result not found")
            except Exception as err:
                print(f"Exception observed while extracting status of NON CA log files is {err}")

        wb = openpyxl.load_workbook(self.excel_file_name)
        sheet = wb.active
        try:
            def apply_thin_borders(sheet):
                for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                    for cell in row:
                        if cell.value:
                            cell.border = self.thin_border
            apply_thin_borders(sheet)

            print(f"Status like for NON Ca logs {self.status_list_non_ca}")
            for file in filtered_log_files:
                file = str(file).split(".")[0]
                file_path_column = [cell.value for cell in sheet['B'] if cell.row > 1]
                if file in file_path_column:
                    row_index = None
                    print(f"All Script Name found {file}")
                    for idx, cell_value in enumerate(file_path_column, start=2):
                        if cell_value == file:
                            row_index = idx
                            break
                    if row_index is not None:
                        self.start_index_list_non_ca.append(row_index)
                        for row, status in zip(self.start_index_list_non_ca, self.status_list_non_ca):
                            if status == "FAIL":
                                for failure in self.error_line_list:
                                    sheet.cell(row=row, column=6, value=status)
                                    sheet.cell(row=row, column=7, value=failure)
                                    sheet.cell(row=row, column=6).font = self.red_font
                                    sheet.cell(row=row, column=7).font = self.red_font
                            elif status == "STOPPED":
                                sheet.cell(row=row, column=6, value=status)
                                sheet.cell(row=row, column=6).font = self.red_font
                            else:
                                sheet.cell(row=row, column =6, value =status)
                        print("Sub scripts are updated")
                else:
                    print(f"{file} not found in column 'B'. Skipping...")

            wb.save(self.excel_file_name)
        except Exception as e:
            print(f"Error: {e}")
        finally:
            wb.close()

    def merging_cells_of_all_scripts(self):
        wb = openpyxl.load_workbook(self.excel_file_name)
        sheet = wb.active
        try:
            for start, end in zip(self.start_index_list, self.len_tc_name_list):
                end_cell = start + end - 1
                self.delete_rows.append(end_cell + 1)
                sheet[f"B{start}"].value = sheet[f"B{end_cell + 1}"].value
            self.delete_rows = list(reversed(self.delete_rows))
            for row_index in self.delete_rows:
                sheet.delete_rows(row_index)

            decrement_pattern = 0
            modified_start_list = []
            modified_end_list = []

            for value in self.start_index_list:
                modified_start_list.append(value - decrement_pattern)
                decrement_pattern += 1

            for start, length in zip(modified_start_list, self.len_tc_name_list):
                modified_end_list.append(start + length -1)

            for row_idx, spec_val, actual_val in zip(modified_start_list, self.specified_subscripts_nos, self.len_tc_name_list):
                if actual_val < spec_val:
                    sheet.cell(row=row_idx, column=3, value=spec_val)
                    sheet.cell(row=row_idx, column=4, value=actual_val)
                    sheet.cell(row=row_idx, column=4).font = self.red_font
                elif actual_val > spec_val:
                    sheet.cell(row=row_idx, column=3, value=spec_val)
                    sheet.cell(row=row_idx, column=4, value=actual_val)
                    sheet.cell(row=row_idx, column=3).font = self.saffron
                else:
                    sheet.cell(row=row_idx, column=3, value=spec_val)
                    sheet.cell(row=row_idx, column=4, value=actual_val)

            for start, end in zip (modified_start_list, modified_end_list):
                sheet.merge_cells(f"B{start}:B{end}")
                sheet.merge_cells(f"C{start}:C{end}")
                sheet.merge_cells(f"D{start}:D{end}")

            wb.save(self.excel_file_name)
        except Exception as e:
            print(f"Error: {e}")
        finally:
            wb.close()

    def updating_excel_with_subscripts_status_failures(self):
        for tc_name in self.testcase_name_list:
            print(tc_name)
        print(f"Total no. of subscripts are {self.length_of_testcase_name}")
        print(self.sub_script_status_list)
        print(f"Total no. of subscripts status are {len(self.sub_script_status_list)}")
        print(self.error_line_list)

        wb = openpyxl.load_workbook(self.excel_file_name)
        sheet = wb.active
        try:
            for txt_file in self.ca_log_names:
                file_path_column = [cell.value for cell in sheet['B'] if cell.row > 1]
                if txt_file not in file_path_column:
                    print(file_path_column)
                    pass
                elif txt_file in file_path_column:
                    row_index = None
                    print(f"All Script Name found {txt_file}")
                    for idx, cell_value in enumerate(file_path_column, start=2):
                        if cell_value == txt_file:
                            row_index = idx
                            break
                    if row_index is not None:
                        self.start_index_list.append(row_index)
                        for testcase_name in reversed(self.testcase_name_list):
                            sheet.insert_rows(row_index, amount=1)
                            sheet.cell(row=row_index, column=5, value=testcase_name)
                        print("Sub scripts are updated")
                else:
                    print(f"{txt_file} not found in column 'B'. Skipping...")

            wb.save(self.excel_file_name)
        except Exception as e:
            print(f"Error: {e}")
        finally:
            wb.close()

    def creating_testcase_name(self):
        for txt_file in self.ca_log_names:
            txt_file_full_path = os.path.join(self.log_directory_path, txt_file + '.log')
            print("\n\r")
            print(f"Reading txt format log file is {txt_file}")
            try:
                with open(txt_file_full_path, 'r') as file:
                    content = file.read()
                    self.testcase_name_list = []
                    self.sub_script_status_list = []
                    self.error_line_list = []
                    search_string = "Started Running script"
                    if search_string in content:
                        lines = content.split('\n')
                        for line_number, line in enumerate(lines, start=1):
                            if search_string in line:
                                testcase_name = line.split("script")[1].strip()
                                self.testcase_name_list.append(testcase_name)
                    else:
                        print(f"No subscripts found")

                    self.testcase_name_list = sorted(self.testcase_name_list)
                    self.length_of_testcase_name = len(self.testcase_name_list)
                    self.len_tc_name_list.append(self.length_of_testcase_name)

                    if pass_pattern in content:
                        self.status_list.append("PASS")
                        print("Main log result is passed")
                        for i in range(1, self.length_of_testcase_name + 1):
                            self.sub_script_status_list.append("PASS")
                        print(self.sub_script_status_list)

                    elif fail_pattern in content:
                        self.status_list.append("FAIL")
                        print("Main log result is Failed")

                        for each_tc_name in self.testcase_name_list:
                            print(f"iterating over list of sub scripts {each_tc_name}")
                            start_pattern = f"Started Running script {each_tc_name}"
                            end_pattern = f"Finished Running script {each_tc_name}"
                            error_pattern = f"Failed Running script .*{re.escape(each_tc_name)}.* with error"
                            if start_pattern and end_pattern in content:
                                self.sub_script_status_list.append("PASS")
                                print(f"{each_tc_name} is Passed")
                            elif start_pattern in content and re.search(error_pattern, content):
                                self.sub_script_status_list.append("FAIL")
                                print(f"{each_tc_name} is Failed")

                                lines = content.split('\n')
                                for i in range(len(lines)):
                                    if re.search(error_pattern, lines[i]):
                                        if i + 2 < len(lines):
                                            error_line = lines[i + 2].strip()
                                            print(f"Next line after 'TestFailedError': {error_line}")
                                            self.error_line_list.append(error_line)
                                        break
                            print(self.sub_script_status_list)

                    elif (pass_pattern not in content) and (fail_pattern not in content):
                        self.status_list.append("STOPPED")
                        print("Main log result is Stopped")
                        for i in range(1, self.length_of_testcase_name + 1):
                            self.sub_script_status_list.append("STOPPED")
                        print(self.sub_script_status_list)

                    wb = openpyxl.load_workbook(self.excel_file_name)
                    sheet = wb.active
                    file_path_column = [cell.value for cell in sheet['B'] if cell.row > 1]
                    if txt_file not in file_path_column:
                        pass
                    elif txt_file in file_path_column:
                        row_index = None
                        print(f"All Script Name found {txt_file}")
                        for idx, cell_value in enumerate(file_path_column, start=2):
                            if cell_value == txt_file:
                                row_index = idx
                                break
                        if row_index is not None:
                            self.start_index_list.append(row_index)
                            for testcase_name, status in zip((reversed(self.testcase_name_list)), self.sub_script_status_list):
                                sheet.insert_rows(row_index, amount=1)
                                sheet.cell(row=row_index, column=5, value=testcase_name)
                                sheet.cell(row=row_index, column=6, value=status)
                                if status == "FAIL":
                                    sheet.cell(row=row_index, column=6).font = self.red_font

                                if status == "FAIL":
                                    print(self.error_line_list)
                                    for failure in self.error_line_list:
                                        sheet.cell(row=row_index, column=7, value=failure)
                                        sheet.cell(row=row_index, column=7).font = self.red_font
                                if status == "STOPPED":
                                    sheet.cell(row=row_index, column=6).font = self.red_font

                            print("Sub scripts are updated")
                    wb.save(self.excel_file_name)

            except Exception as err:
                print(f"Exception occurred during validating Pass/Fail is {err}")

    def creating_allscript_name(self):
        try:
            column_num = 2
            current_row = 2
            wb = openpyxl.load_workbook(self.excel_file_name)
            sheet = wb.active

            for log_name in self.log_files_list:
                current_col = openpyxl.utils.get_column_letter(column_num)
                log_name_value = log_name.split(".")[0]
                if log_name.startswith('CA'):
                    self.ca_log_names.append(log_name_value)
                sheet[f"{current_col}{current_row}"].value = log_name_value
                current_row += 1
            wb.save(self.excel_file_name)
            print("Successfully update the log file names as ALLScripts Name ")
        except Exception as err:
            print(f"Error occurred while creating ALLScripts Name  column is {err}")

    def iterating_over_log_directory(self):
        try:
            self.log_files_list = os.listdir(self.log_directory_path)
            self.log_files_list = [file for file in self.log_files_list if file.endswith('.log')]
            self.log_files_list.sort(key=lambda x: not x.startswith("CA"))
            print(self.log_files_list)
            self.log_list_len = len(self.log_files_list)
            print(f"List of Log files are {self.log_files_list}")
            print(f"Total no. of log files are  {self.log_list_len}")
        except FileNotFoundError:
            print("File are not found or provided path is incorrect")
        except Exception as err:
            print(f"Error occurred while iteration log files is {err}")

    def creating_CQ_Parser_excel_template(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            sheet['A1'] = "SL No."
            sheet['A1'].font = Font(bold=True)
            sheet['B1'] = "CallAllScript Name"
            sheet['B1'].font = Font(bold=True)

            sheet['C1'] = "Specified No. of TCs"
            sheet['C1'].font = Font(bold=True)
            sheet['D1'] = "Actual No. of TCs"
            sheet['D1'].font = Font(bold=True)

            sheet['E1'] = "TestCase Name"
            sheet['E1'].font = Font(bold=True)
            sheet['F1'] = "Status"
            sheet['F1'].font = Font(bold=True)
            sheet['G1'] = "Failure Details"
            sheet['G1'].font = Font(bold=True)

            workbook.save(self.excel_file_name)
            print("Successfully created the CQ Compliance Excel Template")

        except Exception as err:
            print(f"Error observed while creating the {self.excel_file_name} file")

def input_from_cmd_line():
    print("=" * 55 + "CQ Parsing Started" + "=" * 55)
    print("File name for generating CQ Parser file is : CQ_Parser")
    excel_file_name = input("Enter the excel file name: ")
    log_directory_path = input("Enter the path of log files: ")
    return excel_file_name, log_directory_path


if __name__ == "__main__":
    excel_file_name, log_directory_path = input_from_cmd_line()
    print(f"Entered CQ Parser file is {excel_file_name}")
    print(f"Entered log files directory is {log_directory_path}")
    CQ = CQParser(excel_file_name, log_directory_path)
    CQ.creating_CQ_Parser_excel_template()
    CQ.iterating_over_log_directory()
    CQ.creating_allscript_name()
    CQ.creating_testcase_name()
    CQ.merging_cells_of_all_scripts()
    CQ.extracting_status_for_non_CA_logs()
    CQ.excel_styling()