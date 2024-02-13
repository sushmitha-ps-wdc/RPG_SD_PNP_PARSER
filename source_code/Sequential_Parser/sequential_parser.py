"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:15-11-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment, Border, Side
from bs4 import BeautifulSoup
import pandas as pd

class PerformanceParser:

    def __init__(self, template, path_for_log):
        self.parser_template = template
        self.log_directory_path = path_for_log
        self.excel_filename = str()
        self.excel1_start_row_index = 4
        self.excel2_start_row_index = 5
        self.excel3_start_row_index = 5
        self.excel_file_path = os.path.join(self.log_directory_path, self.parser_template + '.xlsx')
        self.sd_speed_list = []
        self.log_list = []
        self.log_list_length = int()
        self.blue_colour = PatternFill(start_color="E7ECFD", end_color="E7ECFD", fill_type="solid")
        self.orange_colour = PatternFill(start_color="FFD27F", end_color="FFD27F", fill_type="solid")
        self.red_colour = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.start_row_index = int()
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        self.speed_directory_extension = ['25MHz', '50MHz', 'SDR50', 'SDR104']
        self.speed_file_path = str()
        self.log_files_list = []
        self.string_to_find = "Throughput"
        self.fill_zero_value_column = 2
        self.average_cell_index = str()
        self.prefill_cell_index_25MHz = str()
        self.value_indices = []
        self.row_index_start = int()

    @staticmethod
    def extract_values(soup, string_to_find):
        """
        Function to extract the values from html file by search "Throughput" and split by "=" operator
        """
        values = []
        for tag in soup.find_all(string=lambda x: x and string_to_find in x):
            value = tag.split('=')[-1].strip()
            values.append(value)
        return values

    def excel_styling(self):
        """
        Definition to excel styling such as applying thin border, adjust cells width,
        colouring, bold font, merging cells
        """
        def adjust_cell_width(sheet): # To adjust column width
            max_column_length = 0
            for col in sheet.columns:
                if col[0].column != 1:
                    max_length = max(len(str(cell.value)) for cell in col)
                    max_column_length = max(max_length, max_column_length)
            for col in sheet.columns:
                column = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[column].width = 10 if col[0].column == 1 else max_column_length * 0.7

        def apply_thin_borders(sheet):
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = self.thin_border

        def apply_centre_alignment(sheet):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        def merge_cells(sheet, cells_to_merge):
            for cells in cells_to_merge:
                sheet.merge_cells(cells)

        def bold_font_cells(sheet,bold_cells):
            for cells in bold_cells:
                sheet[cells].font = openpyxl.styles.Font(bold=True)

        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)
            sheet = workbook.active
            if self.parser_template == "Sequential":
                apply_centre_alignment(sheet)
                apply_thin_borders(sheet)
                adjust_cell_width(sheet)
                merge_cells(sheet, ['B2:C2', 'D2:E2', 'F2:G2', 'H2:I2', 'A1:I1', 'A2:A3'])

                bold_font_cells(sheet,['B3', 'D3', 'F3', 'H3', 'C3', 'E3', 'G3', 'I3', 'B2', 'D2', 'F2', 'H2', 'A1', 'A2'])

                sheet['A1'].fill = self.orange_colour
                for row in sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=sheet.max_column):
                    for cell in row:
                        cell.fill = self.blue_colour
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.fill = self.blue_colour

            elif self.parser_template == "Sequential_both_iterations":
                merge_cells(sheet,['B3:C3', 'F3:G3', 'J3:K3', 'N3:O3', 'D3:E3', 'H3:I3', 'L3:M3', 'P3:Q3', 'B2:E2',
                                     'A1:Q1','A2:A4', 'F2:I2', 'J2:M2', 'N2:Q2'])
                apply_centre_alignment(sheet)
                apply_thin_borders(sheet)
                adjust_cell_width(sheet)

                bold_font_cells(sheet,['A2', 'B2', 'F2', 'J2', 'N2', 'B3', 'D3', 'F3', 'H3', 'J3', 'L3', 'N3', 'P3', 'B4'])
                bold_font_cells(sheet,['C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'O4', 'P4', 'Q4'])

                sheet['A1'].fill = self.orange_colour
                for row in sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=sheet.max_column):
                    for cell in row:
                        cell.fill = self.blue_colour
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.fill = self.blue_colour

            elif self.parser_template == "Sequential_with_prefill_data":
                merge_cells(sheet, ['A2:A4', 'B2:F2', 'G2:K2', 'L2:P2', 'Q2:U2', 'C3:D3', 'H3:I3', 'M3:N3', 'R3:S3',
                'E3:F3', 'J3:K3', 'O3:P3', 'T3:U3', 'A1:U1'])
                apply_centre_alignment(sheet)
                apply_thin_borders(sheet)
                adjust_cell_width(sheet)

                try:
                    bold_font_cells(sheet, ['A1', 'A2', 'B2', 'G2', 'L2', 'Q2', 'B3', 'G3', 'L3', 'Q3', 'C3', 'H3', 'M3'
                        , 'R3', 'E3', 'J3', 'O3', 'T3'])
                    bold_wr_cells = []
                    # created loop for iterating only from 'B' to 'U' (B4 TO U4) and make text Bold
                    for i in range(ord('B'), ord('V')):
                        col_letter = chr(i)
                        bold_wr_cells.append(f"{col_letter}4")
                    for cell in bold_wr_cells:
                        sheet[cell].font = openpyxl.styles.Font(bold=True)

                except Exception as err:
                    print(f"Error while making bold text for column names {err}")

                sheet['A1'].fill = self.orange_colour
                for row in sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=sheet.max_column):
                    for cell in row:
                        cell.fill = self.blue_colour
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.fill = self.blue_colour

                for row in range(self.excel3_start_row_index, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=2)
                    if cell.value == 0:  # Check if the value is specifically 0
                        cell.border = self.thin_border

            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if cell.value is None:
                        cell.fill = self.red_colour

            workbook.save(self.excel_file_path)
            print(f"Successfully updated excel sheet frames design for {self.excel_file_path}")
        except Exception as err:
            print(f"Exception occurred while updating excel sheet frames design for {self.excel_file_path}: {err}")

    def calculate_and_update_average(self):
        """
        1. Definition to calculate average value of data in respective columns and update in succeeded empty row.
        2. Creating the "Average" text inside Sample No. column in succeeded empty row with fixed column 'A'
        """
        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)
            sheet = workbook.active

            if self.parser_template == "Sequential":
                self.start_row_index = self.excel1_start_row_index
            elif self.parser_template == "Sequential_both_iterations":
                self.start_row_index = self.excel2_start_row_index
            elif self.parser_template == "Sequential_with_prefill_data":
                self.start_row_index = self.excel2_start_row_index

            end_row = sheet.max_row - 1
            last_column = sheet.max_column
            average_formula_row = self.start_row_index + self.log_list_length

            for col_num in range(2, last_column + 1):
                col_lett = openpyxl.utils.get_column_letter(col_num)
                avg_formula = f"=ROUND(AVERAGE({col_lett}{self.start_row_index}:{col_lett}{end_row}), 3)"

                cell = sheet.cell(row=average_formula_row , column=col_num, value=avg_formula)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            workbook.save(self.excel_file_path)
        except Exception as err:
            print(f"Error occurred at updating average values is {err}")

    def sequential_excel_perf_values_update(self):
        """
        Definition to iterate over required Throughput values and update in Sequential.xlsx file.
        NOTE: 2nd Iteration write and read values are considered for updating excel file
        """
        for file_name in self.speed_directory_extension:
            self.row_index_start = 2
            if file_name == "25MHz":
                self.value_indices = [2, 3]
                column_letter = 1
            elif file_name == "50MHz":
                self.value_indices = [3, 4]
                column_letter = 3
            elif file_name == "SDR50":
                self.value_indices = [3, 4]
                column_letter = 5
            elif file_name == "SDR104":
                self.value_indices = [3, 4]
                column_letter = 7

            speed_file_path = os.path.join(self.log_directory_path, file_name)
            print(f"Data rate folder path {speed_file_path}")

            log_files_list = os.listdir(speed_file_path)
            log_files_list = sorted(log_files_list, key=lambda x: int(x.split('.')[0]))

            for html_file_name in log_files_list:
                print(f"current html file {html_file_name}")
                full_log_file_path = os.path.join(speed_file_path, html_file_name)
                print(f"complete html file path {full_log_file_path}")
                with open(full_log_file_path, 'r') as file:
                    html_content = file.read()

                    soup = BeautifulSoup(html_content, 'html.parser')
                    values_dict = self.extract_values(soup, self.string_to_find)
                    print(values_dict)
                    df = pd.DataFrame()
                    try:
                        df = pd.read_excel(self.excel_file_path)
                        print("DataFrame shape:", df.shape)
                        required_values_list = []
                        for ind in self.value_indices:
                            value = values_dict[ind]
                            if value == "":
                                required_values_list.append(str(""))
                            if value is not None and value != "":
                                required_values_list.append(round(float(str(value).split("(")[0]), 3))
                        print(f"Current fetched value is {required_values_list}")
                        df.iloc[self.row_index_start, column_letter] = required_values_list[0]
                        column_letter = column_letter + 1
                        df.iloc[self.row_index_start, column_letter] = required_values_list[1]
                        column_letter = column_letter - 1
                    except Exception as err:
                        print(f"{err}")
                    finally:
                        df.to_excel(self.excel_file_path, index=False)
                        print("Excel file updated successfully.")
                self.row_index_start = self.row_index_start + 1


    def sequential_both_iterations_excel_perf_values_update(self):
        """
        Definition to iterate over required Throughput values and update in Sequential_both_iterations.xlsx file.
        NOTE: 25MHz Four occurence of "Throughput values" are considered as 1st iteration (W,R) and 2nd Iteration (W,R)
              For 50MHz,SDR50 and SDR104, omitted the 1st "Throughput value" as prefilled value.
        """
        for file_name in self.speed_directory_extension:
            self.row_index_start = 3
            if file_name == "25MHz":
                self.value_indices = [0, 1, 2, 3]
                column_letter = 1
            elif file_name == "50MHz":
                self.value_indices = [1, 2, 3, 4]
                column_letter = 5
            elif file_name == "SDR50":
                self.value_indices = [1, 2, 3, 4]
                column_letter = 9
            elif file_name == "SDR104":
                self.value_indices = [1, 2, 3, 4]
                column_letter = 13

            speed_file_path = os.path.join(self.log_directory_path, file_name)
            print(f"Data rate folder path {speed_file_path}")

            log_files_list = os.listdir(speed_file_path)
            log_files_list = sorted(log_files_list, key=lambda x: int(x.split('.')[0]))

            for html_file_name in log_files_list:
                full_log_file_path = os.path.join(speed_file_path, html_file_name)
                print(f"complete html file path {full_log_file_path}")
                with open(full_log_file_path, 'r') as file:
                    html_content = file.read()

                    soup = BeautifulSoup(html_content, 'html.parser')
                    string_to_find = "Throughput"
                    values_dict = self.extract_values(soup, string_to_find)
                    print(values_dict)
                    try:
                        df = pd.read_excel(self.excel_file_path)
                        required_values_list = []
                        for ind in self.value_indices:
                            value = values_dict[ind]
                            if value == "":
                                required_values_list.append(str(""))
                            if value is not None and value != "":
                                required_values_list.append(round(float(str(value).split("(")[0]), 3))
                        print(f"Current fetched value is {required_values_list}")
                        for i, value in enumerate(required_values_list):
                            df.iloc[self.row_index_start, column_letter] = value
                            column_letter += 1
                            if i == 3:
                                column_letter -= 4
                    except Exception as err:
                        print(f"{err}")
                    finally:
                        df.to_excel(self.excel_file_path, index=False)
                        print("Excel file updated successfully.")
                self.row_index_start = self.row_index_start + 1

    def sequential_prefill_with_data_excel_perf_values_update(self):
        """
        Definition to iterate over required Throughput values and update in Sequential_both_iterations.xlsx file.
        NOTE: 25MHz Four occurrence of "Throughput values" are considered as 1st iteration (W,R) and 2nd Iteration (W,R)
             For 25MHz prefill will be 0 (Zero)
             For 50MHz,SDR50 and SDR104, five occurrence of "Throughput value" are considered as prefilled value,
              1st iteration(W,R) and 2nd Iteration(W,R)
        """

        for file_name in self.speed_directory_extension:
            self.row_index_start = 3
            if file_name == "25MHz":
                self.value_indices = [0, 1, 2, 3]
                column_letter = 2
            elif file_name == "50MHz":
                self.value_indices = [0,1, 2, 3, 4]
                column_letter = 6
            elif file_name == "SDR50":
                self.value_indices = [0,1, 2, 3, 4]
                column_letter = 11
            elif file_name == "SDR104":
                self.value_indices = [0,1, 2, 3, 4]
                column_letter = 16

            speed_file_path = os.path.join(self.log_directory_path, file_name)
            print(f"Data rate folder path {speed_file_path}")

            log_files_list = os.listdir(speed_file_path)
            log_files_list = sorted(log_files_list, key=lambda x: int(x.split('.')[0]))

            for html_file_name in log_files_list:
                print(f"current html file {html_file_name}")
                full_log_file_path = os.path.join(speed_file_path, html_file_name)
                print(f"complete html file path {full_log_file_path}")
                with open(full_log_file_path, 'r') as file:
                    html_content = file.read()

                    soup = BeautifulSoup(html_content, 'html.parser')
                    string_to_find = "Throughput"
                    values_dict = self.extract_values(soup, string_to_find)
                    print(values_dict)
                    try:
                        df = pd.read_excel(self.excel_file_path)
                        print("DataFrame shape:", df.shape)
                        if file_name == "25MHz":
                            required_values_list = []
                            for ind in self.value_indices:
                                value = values_dict[ind]
                                if value == "":
                                    required_values_list.append(str(""))
                                if value is not None and value != "":
                                    required_values_list.append(round(float(str(value).split("(")[0]), 3))
                            for i, value in enumerate(required_values_list):
                                df.iloc[self.row_index_start, column_letter] = value
                                column_letter += 1
                                if i == 3:
                                    column_letter -= 4
                        else:
                            required_values_list = []
                            for ind in self.value_indices:
                                value = values_dict[ind]
                                if value == "":
                                    required_values_list.append(str(""))
                                if value is not None and value != "":
                                    required_values_list.append(round(float(str(value).split("(")[0]), 3))
                            for i, value in enumerate(required_values_list):
                                df.iloc[self.row_index_start, column_letter] = value
                                column_letter += 1
                                if i == 4:
                                    column_letter -= 5
                    except Exception as err:
                        print(f"{err}")
                    finally:
                        df.to_excel(self.excel_file_path, index=False)
                        print("Excel file updated successfully.")
                self.row_index_start = self.row_index_start + 1

    def update_excel_with_performance_values(self):
        try:
            if self.parser_template == "Sequential":
                self.sequential_excel_perf_values_update()
            elif self.parser_template == "Sequential_both_iterations":
                self.sequential_both_iterations_excel_perf_values_update()
            elif self.parser_template == "Sequential_with_prefill_data":
                self.sequential_prefill_with_data_excel_perf_values_update()
        except Exception as err:
            print(f"Error occurred while updating performance values")

    def update_sample_num_column(self):
        """
        Definition to update the No. of Sample No. under Column 'A' and fill with respective log file name.
        """
        try:
            if self.parser_template == "Sequential":
                self.start_row_index = self.excel1_start_row_index
            elif self.parser_template == "Sequential_both_iterations":
                self.start_row_index = self.excel2_start_row_index
            elif self.parser_template == "Sequential_with_prefill_data":
                self.start_row_index = self.excel3_start_row_index

            wb = openpyxl.load_workbook(self.excel_file_path)
            sheet = wb.active
            for volume, name in zip((range(self.start_row_index, self.start_row_index + self.log_list_length)), self.log_list):
                row_name = int(name.split(".")[0])
                sheet["A" + f"{volume}"].value = row_name
                sheet["A" + f"{volume}"].font = Font(bold=True)

            average_row_index = self.start_row_index + int(self.log_list_length)
            self.average_cell_index = str('A' + str(average_row_index))
            sheet[self.average_cell_index].value = "Average"
            sheet[self.average_cell_index].font = Font(bold=True)
            sheet[self.average_cell_index].border = self.thin_border
            sheet[self.average_cell_index].fill = self.blue_colour
            print(f"Text 'Average' updated in {self.average_cell_index}")

            if self.parser_template == "Sequential_with_prefill_data":
                for volume in range(self.start_row_index, self.start_row_index + self.log_list_length):
                    pre_fill_value_for_25MHz = int(0)
                    self.prefill_cell_index_25MHz = str("B" + f"{volume}")
                    sheet[self.prefill_cell_index_25MHz].value = pre_fill_value_for_25MHz
                print(f"Pre fill data for 25MHz has been updated with 0 int value")

            wb.save(self.excel_file_path)
            print(f"{self.log_list_length} samples created from row '{self.start_row_index}' in the Excel file.")

        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as e:
            print(f"An error occurred during updating sample number column: {e}")

    def iterating_over_log_directory(self):
        """
        Defintion to get the list of log_files_names and total no. of log file
        """
        try:
            path = os.path.join(self.log_directory_path, "25MHz")
            self.sd_speed_list = os.listdir(path)
            self.log_list = [file for file in self.sd_speed_list if file.endswith('htm')]
            self.log_list = sorted(self.log_list, key=lambda x: int(x.split('.')[0]))
            self.log_list_length = len(self.log_list)
            print(f"Total number of log files in a directory: {self.log_list_length}")
        except FileNotFoundError:
            print("File not found or paths are incorrect")
        except Exception as err:
            print(f"Occurred exception while iterating log directory is {err}")

    def create_specified_excel_structure(self):
        """
        Definition to create basic specified excel sheets header
        :return:
        """
        if self.parser_template == "Sequential":
            self.excel_filename = "Sequential.xlsx"
            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['B3'], sheet['D3'], sheet['F3'], sheet['H3'] = ["Write (MB/s)"] * 4
                sheet['C3'], sheet['E3'], sheet['G3'], sheet['I3'] = ["Read (MB/s)"] * 4
                sheet['B2'] = "Low Speed (25 MHz)"
                sheet['C2'], sheet['E2'], sheet['G2'], sheet['I2'], sheet['I2'], sheet['A2'] = [""] * 6
                sheet['D2'] = "High Speed (50 MHz)"
                sheet['F2'] = "SDR50 (100 MHz)"
                sheet['H2'] = "SDR104 (208 MHz)"
                sheet['A1'] = "Sequential Performance"
                sheet['A2'] = "Sample No."
                workbook.save(self.excel_file_path)
                print("successfully created header for Sequential Excel file")
            except Exception as err:
                print(f"Exception occurred at creating header for Sequential Excel file {err}")

        elif self.parser_template == "Sequential_both_iterations":
            self.excel_filename = "Sequential_both_iterations.xlsx"
            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['B4'], sheet['D4'], sheet['F4'], sheet['H4'], sheet["J4"], sheet["L4"], sheet["N4"], sheet[
                    "P4"] = ["Write (MB/s)"] * 8
                sheet['C4'], sheet['E4'], sheet['G4'], sheet['I4'], sheet["K4"], sheet["M4"], sheet["O4"], sheet[
                    "Q4"] = ["Read (MB/s)"] * 8
                sheet['B3'], sheet['F3'], sheet['J3'], sheet['N3'] = ["1st Iteration Perf"] * 4
                sheet['D3'], sheet['H3'], sheet['L3'], sheet['P3'] = ["2nd Iteration Perf"] * 4
                sheet['B2'] = "Low Speed(25MHz)"
                sheet['F2'] = "High Speed(50MHz)"
                sheet['J2'] = "SDR50(100MHz)"
                sheet['N2'] = "SDR104(208MHz)"
                sheet['A1'] = "Sequential both iterations Performance"
                sheet['A2'] = "Sample No."

                workbook.save(self.excel_file_path)
                print("successfully created header for Sequential both Iterations Excel file")
            except Exception as err:
                print(f"Exception occurred at creating header for Sequential both Iterations Excel file {err}")
        elif self.parser_template == "Sequential_with_prefill_data":
            self.excel_filename = "Sequential_with_prefill_data.xlsx"
            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['B4'], sheet['C4'], sheet['E4'], sheet['G4'], sheet['H4'], sheet['J4'], sheet['L4'], sheet[
                    'M4'], sheet['O4'], sheet['Q4'], sheet['R4'], sheet['T4'] = ["Write (MB/s)"] * 12
                sheet['D4'], sheet['F4'], sheet['I4'], sheet['K4'], sheet['N4'], sheet['P4'], sheet['S4'], sheet[
                    'U4'] = ["Read (MB/s)"] * 8
                sheet['B3'], sheet['G3'], sheet['L3'], sheet['Q3'] = ["Filling Perf"] * 4
                sheet['C3'], sheet['H3'], sheet['M3'], sheet['R3'] = ["1st iteration Perf"] * 4
                sheet['E3'], sheet['J3'], sheet['O3'], sheet['T3'] = ["2nd iteration Perf"] * 4
                sheet['B2'] = "Low speed(25 MHz)"
                sheet['G2'] = "High speed(50 MHz)"
                sheet['L2'] = "SDR50(100 MHz)"
                sheet['Q2'] = "SDR104(208 MHz)"
                sheet['A1'] = "Sequence Performance with Prefill Data"
                sheet['A2'] = "Sample No."

                workbook.save(self.excel_file_path)
                print("successfully created header for Sequential both Iterations Excel file")
            except Exception as err:
                print(f"Exception occurred at creating header for Sequential both Iterations Excel file {err}")

def inputs_from_cmd_line():
    print("=" * 55 + "Command line inputs for Sequential Parsing:" + "=" * 55 + "\n")
    print("Choose the below excel sheets to generate:")
    print("\n Sequential\n Sequential_both_iterations\n Sequential_with_prefill_data\n")
    parsing_template_input = input("Enter the required parsing template: ")
    path_for_log_files_input = input("Enter the log files directory path: ")
    print("Capturing inputs for Performance parsing are done. Below are the inputs for parsing")
    return parsing_template_input, path_for_log_files_input

if __name__ == "__main__":
    print("=" * 55 + "Started Sequential Parsing" + "=" * 55 + "\n")
    parsing_template, path_for_log_files = inputs_from_cmd_line()
    print("\n\r")
    print(f"Entered parsing template is: {parsing_template}")
    print(f"Entered log files directory path is: {path_for_log_files}")
    print("\n\r")
    pp = PerformanceParser(parsing_template, path_for_log_files)
    pp.create_specified_excel_structure()
    pp.iterating_over_log_directory()
    pp.update_sample_num_column()
    pp.update_excel_with_performance_values()
    pp.calculate_and_update_average()
    pp.excel_styling()
