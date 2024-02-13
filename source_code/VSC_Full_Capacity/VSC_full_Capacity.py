"""
WESTERN DIGITAL CORPORATION
Copyright Western Digital Corporation All Rights Reserved.

The sourcecode contained or described here in and all documents related to
the sourcecode("Material") are owned by WesternDigital Corporation or its suppliers
or licensors. Title to the Material remains with Western Digital Corporation or its
suppliers and licensors.

No license under any patent,copyright,trade secret or other intellectual
property right is granted to or conferred upon you by disclosure or delivery
of the Materials,either expressly,by implication,inducement,estoppel or
otherwise.Any license under such intellectual property rights must be express
and approved by WesternDigital in writing.
###############################################################################################

#InitialVersion:1.0
#Date:13-12-2023
#Author(s):Sushmitha P.S

##################################################################################################
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from bs4 import BeautifulSoup
import shutil
import pandas as pd
from openpyxl.chart import LineChart, Reference

class VSCFullCapacity:

    def __init__(self,excel_file_name, log_directory_path):
        self.file_name = excel_file_name
        self.log_directory_path = log_directory_path
        self.excel_file_name = os.path.join(self.log_directory_path, self.file_name + ".xlsx")
        self.working_directory = os.getcwd()
        self.initial_excel_file = os.path.join(self.working_directory, self.file_name + ".xlsx")
        self.items_to_search = ['VSC6', 'VSC10', 'VSC30', 'VSC60', 'VSC90']
        self.found_items = []
        self.log_files = str()
        self.html_file_path = str()
        self.blue_colour = PatternFill(start_color="E7ECFD", end_color="E7ECFD", fill_type="solid")
        self.pw_performance_values = []
        self.length_found_items = int()
        self.spec_val = int()
        self.red_font = Font(color='00FF0000', bold=True)
        self.red_fill = PatternFill(start_color="FFFF0000",end_color="FFFF0000", fill_type="solid")


    def graph_design(self, sheet):
        chart = LineChart()
        chart.title = "VSC Full Capacity"
        chart.x_axis.title = 'AU Number'
        chart.y_axis.title = 'Performance [MB/s]'
        chart.legend.position = 'b'

        for col_num in range(1, sheet.max_column + 1):
            data = [cell[0].value for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_num, max_col=col_num)]

            hard_data_numeric = [float(value) for value in data]
            data_ref = Reference(sheet, min_col=col_num, min_row=1, max_col=col_num, max_row=len(hard_data_numeric) + 1)
            chart.add_data(data_ref, titles_from_data=True)

        sheet.add_chart(chart, "G2")
        chart.width = 23
        chart.height = 9

    def graph_update_in_excel_sheet(self):
        try:
            workbook = openpyxl.load_workbook(self.excel_file_name)
            sheet = workbook.active
            self.graph_design(sheet)
            workbook.save(self.excel_file_name)

            print("Successfully created the Graphical representation of Performance Values")
        except Exception as err:
            print(f"Error observed: {err}")


    def updating_values_to_excel_sheet(self):
        pw_values_len = len(self.pw_performance_values)
        print(f"Length of pw_values found is {pw_values_len}")

        division_result = pw_values_len // self.length_found_items
        print(f"Total number of pw_values for each VSC is {division_result}")

        vsc_lists = [f'{item}' for item in self.found_items]
        print(vsc_lists)

        created_lists = {}
        for i, vsc_list in enumerate(vsc_lists):
            start_index = i * division_result
            end_index = (i + 1) * division_result
            created_lists[vsc_list] = [float(value) for value in self.pw_performance_values[start_index:end_index]]

        for vsc_list, values in created_lists.items():
            print(f"{vsc_list}: {values}")

        try:
            workbook = openpyxl.load_workbook(self.excel_file_name)
            sheet = workbook.active
            for col_index, column_values in enumerate(created_lists.values(), start=1):
                if col_index == 1:
                    self.spec_val = 6.0
                elif col_index == 2:
                    self.spec_val = 10.0
                elif col_index == 3:
                    self.spec_val = 30.0
                elif col_index == 4:
                    self.spec_val = 60.0
                elif col_index == 5:
                    self.spec_val = 90.0

                for row_index, value in enumerate(column_values, start=2):
                    if value == float(0.0):
                        sheet.cell(row=row_index, column=col_index).font = self.red_font
                    elif value is None or value == " " or value == ' ':
                        sheet.cell(row=row_index, column=col_index).font = self.red_fill
                    elif value < float(self.spec_val):
                        sheet.cell(row=row_index, column=col_index, value=value)
                        sheet.cell(row=row_index, column=col_index).font = self.red_font
                    sheet.cell(row=row_index, column=col_index, value=value)

            workbook.save(self.excel_file_name)
            print("Excel file updated successfully")
            workbook.close()
        except Exception as err:
            print(f"Exception occurred while saving excel file with values is : {err}")

    def extracting_values_from_log_files(self):
        try:
            self.html_file_path = os.path.join(self.log_directory_path, self.log_files)
            print(self.html_file_path)
            with open(self.html_file_path, 'r') as file:
                html_content = file.read()
                soup = BeautifulSoup(html_content, 'html.parser')

                for label_tag in soup.find_all('span', class_='us', string='VSC Pw [MB/s]'):
                    div_tag = label_tag.find_parent('div')
                    if div_tag:
                        next_div_tag = div_tag.find_next('div')
                        print(f"Print parent div: {div_tag}")
                        print(f"Next child div tag: {next_div_tag}")
                        if next_div_tag:
                            value_tag = next_div_tag.find('span', class_='us')
                            if value_tag:
                                int_value = value_tag.text.strip()
                                if int_value is not None and int_value != "":
                                    try:
                                        int_value = round(float(int_value), 8)
                                        print(int_value)
                                        self.pw_performance_values.append(int_value)
                                    except ValueError:
                                        print(f"Error: Unable to convert '{int_value}' to float.")
                                        self.pw_performance_values.append(float(0))
                                else:
                                    self.pw_performance_values.append(float(0))
                            else:
                                self.pw_performance_values.append(float(0))
                                print("No values found")
            file.close()
            print("Successfully fetched all the PW Values from log file")
        except Exception as err:
            print(f"Exception occurred while fetching PW values from log file is {err}")

    def creating_excel_header_as_per_VSC_strings_found(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "VSC Full Capacity-Write"

            for col_num, header in enumerate(self.found_items, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = self.blue_colour
        
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.value = None
            workbook.save(self.excel_file_name)
            print("Successfully created the VSC_Full_Capacity Excel Template")
        except Exception as err:
            print(f"Error observed while creating the {self.excel_file_name} file is {err}")

    def finding_VSC_strings_in_log_file(self):
        try:
            self.log_files = os.listdir(self.log_directory_path)
            print(self.log_files)
            for file in self.log_files:
                if file.endswith('.htm'):
                    print(file)
                    self.log_files = file
                    self.html_file_path = os.path.join(self.log_directory_path, self.log_files)
                    print(self.html_file_path)
                    with open(self.html_file_path, 'r') as f:
                        self.found_items = []
                        html_content = f.read()
                        for item in self.items_to_search:
                            if item in html_content:
                                self.found_items.append(item)
                        print(f" {self.log_files} found VSC items are {self.found_items}")
            self.length_found_items = len(self.found_items)
        except FileNotFoundError:
            print("File are not found or provided path is incorrect")
        except Exception as err:
            print(f"Error occurred while searching VSC Strings in log file is {err}")



def input_from_cmd_line():
    print("=" * 55 + "VSC Full Capacity Parsing Started" + "=" * 55)
    print("File name for generating VSC Full Capacity file is : VSC_Full_Capacity")
    excel_file_name = input("Enter the excel file name: ")
    log_directory_path = input("Enter the path of log files: ")
    return excel_file_name, log_directory_path

if __name__ == "__main__":
    excel_file_name, log_directory_path = input_from_cmd_line()
    print(f"Entered VSC Full Capacity is {excel_file_name}")
    print(f"Entered log files directory is {log_directory_path}")
    VSC_FC = VSCFullCapacity(excel_file_name, log_directory_path)
    VSC_FC.finding_VSC_strings_in_log_file()
    VSC_FC.creating_excel_header_as_per_VSC_strings_found()
    VSC_FC.extracting_values_from_log_files()
    VSC_FC.updating_values_to_excel_sheet()
    VSC_FC.graph_update_in_excel_sheet()

